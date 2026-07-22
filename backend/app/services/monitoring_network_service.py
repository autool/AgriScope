"""田间监测网络、设备故障和病虫害预警业务服务。"""

import asyncio
import hashlib
import io
import json
import mimetypes
import os
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.exceptions import NotFoundException, ValidationException
from app.dao.monitoring_network_dao import (
    AlertRecord,
    AssessmentRecord,
    ConsultationRecord,
    DeviceRecord,
    FaultRecord,
    MonitoringNetworkDAO,
    ReportAssessmentRecord,
    TelemetryRecord,
)
from app.models.monitoring_network import (
    DeviceFault,
    DeviceTelemetry,
    ExpertConsultation,
    MonitoringDevice,
    MonitoringEvent,
    MonitoringStation,
    PestAlert,
    PestAssessment,
    PestModelVersion,
    PestReport,
    PestReportItem,
)
from app.models.workbench import MonitoringProject, ProjectUser
from app.schemas.monitoring_network import (
    AlertCreateRequest,
    AlertDeliverRequest,
    AlertResponse,
    AssessmentCreateRequest,
    AssessmentResponse,
    AssessmentReviewRequest,
    DeviceCreateRequest,
    DeviceResponse,
    ExpertConsultationAnswerRequest,
    ExpertConsultationCreateRequest,
    ExpertConsultationResponse,
    FaultCreateRequest,
    FaultResolveRequest,
    FaultResponse,
    MonitoringEventResponse,
    MonitoringOverviewResponse,
    PestModelCreateRequest,
    PestModelResponse,
    PestReportCreateRequest,
    PestReportItemResponse,
    PestReportResponse,
    PestReportReviewRequest,
    PestReportReviseRequest,
    PestReportSubmitRequest,
    StationCreateRequest,
    StationResponse,
    TelemetryCreateRequest,
    TelemetryResponse,
)
from app.services.project_user_service import ProjectUserService


@dataclass(frozen=True)
class StoredEvidence:
    """已写入受控目录的专家会商证据。"""

    path: Path
    uri: str
    filename: str
    file_size_bytes: int
    checksum_sha256: str
    created_new: bool


@dataclass(frozen=True)
class ReportDownload:
    """校验通过的病虫害报告下载信息。"""

    path: Path
    filename: str
    media_type: str
    checksum_sha256: str


class MonitoringNetworkService:
    """编排监测资源、遥测、故障、模型复核和告警闭环。"""

    def __init__(
        self,
        dao: MonitoringNetworkDAO | None = None,
        user_service: ProjectUserService | None = None,
        storage_root: Path | None = None,
    ) -> None:
        """初始化田间监测业务服务。

        Args:
            dao: 监测网络 DAO。
            user_service: 项目身份服务。
            storage_root: 病虫害报告和会商证据受控目录。

        Returns:
            None: 无返回值。
        """
        self.dao = dao or MonitoringNetworkDAO()
        self.user_service = user_service or ProjectUserService()
        self.storage_root = (
            storage_root
            or Path(__file__).resolve().parents[2] / "storage" / "monitoring-reports"
        ).resolve()

    async def _require_project(
        self,
        db: AsyncSession,
        project_code: str,
    ) -> MonitoringProject:
        """解析项目上下文。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。

        Returns:
            MonitoringProject: 已存在项目。
        """
        project = await self.dao.get_project_by_code(db, project_code)
        if project is None:
            raise NotFoundException(f"未找到项目 {project_code}")
        return project

    @staticmethod
    def _event(
        project_id: int,
        entity_type: str,
        entity_code: str,
        event_type: str,
        detail: dict,
        user: ProjectUser,
    ) -> MonitoringEvent:
        """构造稳定身份审计事件。

        Args:
            project_id: 项目主键。
            entity_type: 实体类型。
            entity_code: 实体编号。
            event_type: 事件类型。
            detail: 结构化业务证据。
            user: 当前项目用户。

        Returns:
            MonitoringEvent: 未持久化事件。
        """
        return MonitoringEvent(
            project_id=project_id,
            entity_type=entity_type,
            entity_code=entity_code,
            event_type=event_type,
            detail=detail,
            actor=user.display_name,
            actor_code=user.user_code,
            actor_role=user.role_code,
        )

    @staticmethod
    def _station_response(station: MonitoringStation) -> StationResponse:
        """转换监测站响应。

        Args:
            station: 监测站模型。

        Returns:
            StationResponse: 监测站响应。
        """
        return StationResponse(
            station_code=station.station_code,
            station_name=station.station_name,
            province_code=station.province_code,
            province_name=station.province_name,
            city_code=station.city_code,
            city_name=station.city_name,
            district_code=station.district_code,
            district_name=station.district_name,
            longitude=float(station.longitude),
            latitude=float(station.latitude),
            station_type=station.station_type,
            owner_department=station.owner_department,
            source_name=station.source_name,
            source_uri=station.source_uri,
            source_version=station.source_version,
            evidence_uri=station.evidence_uri,
            evidence_size_bytes=station.evidence_size_bytes,
            evidence_sha256=station.evidence_sha256,
            status=station.status,
            registered_by=station.registered_by,
            registered_by_code=station.registered_by_code,
            registered_by_role=station.registered_by_role,
            created_at=station.created_at,
        )

    @staticmethod
    def _device_response(record: DeviceRecord) -> DeviceResponse:
        """转换设备响应。

        Args:
            record: 带站点编号的设备记录。

        Returns:
            DeviceResponse: 设备响应。
        """
        device = record.device
        return DeviceResponse(
            station_code=record.station_code,
            device_code=device.device_code,
            device_name=device.device_name,
            device_type=device.device_type,
            vendor=device.vendor,
            model_number=device.model_number,
            serial_number=device.serial_number,
            owner_department=device.owner_department,
            installed_at=device.installed_at,
            photo_uri=device.photo_uri,
            photo_size_bytes=device.photo_size_bytes,
            photo_sha256=device.photo_sha256,
            status=device.status,
            last_telemetry_at=device.last_telemetry_at,
            registered_by=device.registered_by,
            registered_by_code=device.registered_by_code,
            registered_by_role=device.registered_by_role,
            created_at=device.created_at,
        )

    @staticmethod
    def _telemetry_response(record: TelemetryRecord) -> TelemetryResponse:
        """转换设备遥测响应。

        Args:
            record: 带设备编号的遥测记录。

        Returns:
            TelemetryResponse: 遥测响应。
        """
        telemetry = record.telemetry
        return TelemetryResponse(
            device_code=record.device_code,
            idempotency_key=telemetry.idempotency_key,
            observed_at=telemetry.observed_at,
            metric_code=telemetry.metric_code,
            metric_value=(
                float(telemetry.metric_value)
                if telemetry.metric_value is not None
                else None
            ),
            metric_unit=telemetry.metric_unit,
            payload=telemetry.payload or {},
            evidence_uri=telemetry.evidence_uri,
            evidence_size_bytes=telemetry.evidence_size_bytes,
            evidence_sha256=telemetry.evidence_sha256,
            ingested_by=telemetry.ingested_by,
            ingested_by_code=telemetry.ingested_by_code,
            ingested_by_role=telemetry.ingested_by_role,
            received_at=telemetry.received_at,
        )

    @staticmethod
    def _fault_response(record: FaultRecord) -> FaultResponse:
        """转换设备故障响应。

        Args:
            record: 带设备编号的故障记录。

        Returns:
            FaultResponse: 故障响应。
        """
        fault = record.fault
        return FaultResponse(
            device_code=record.device_code,
            fault_code=fault.fault_code,
            severity=fault.severity,
            reason=fault.reason,
            occurred_at=fault.occurred_at,
            status=fault.status,
            reported_by=fault.reported_by,
            reported_by_code=fault.reported_by_code,
            reported_by_role=fault.reported_by_role,
            resolution_comment=fault.resolution_comment,
            resolution_evidence_uri=fault.resolution_evidence_uri,
            resolved_by=fault.resolved_by,
            resolved_by_code=fault.resolved_by_code,
            resolved_by_role=fault.resolved_by_role,
            resolved_at=fault.resolved_at,
            created_at=fault.created_at,
        )

    @staticmethod
    def _model_response(model: PestModelVersion) -> PestModelResponse:
        """转换病虫害模型版本响应。

        Args:
            model: 模型版本。

        Returns:
            PestModelResponse: 模型响应。
        """
        return PestModelResponse(
            model_code=model.model_code,
            model_version=model.model_version,
            model_name=model.model_name,
            target_type=model.target_type,
            deployment_target=model.deployment_target,
            training_source_uri=model.training_source_uri,
            evaluation_source_uri=model.evaluation_source_uri,
            artifact_uri=model.artifact_uri,
            artifact_size_bytes=model.artifact_size_bytes,
            artifact_sha256=model.artifact_sha256,
            accuracy=float(model.accuracy),
            recall=float(model.recall),
            f1_score=float(model.f1_score),
            roc_auc=float(model.roc_auc),
            status=model.status,
            superseded_by_version=model.superseded_by_version,
            registered_by=model.registered_by,
            registered_by_code=model.registered_by_code,
            registered_by_role=model.registered_by_role,
            created_at=model.created_at,
        )

    @staticmethod
    def _assessment_response(record: AssessmentRecord) -> AssessmentResponse:
        """转换识别结果响应。

        Args:
            record: 带设备和模型身份的识别记录。

        Returns:
            AssessmentResponse: 识别与复核响应。
        """
        assessment = record.assessment
        return AssessmentResponse(
            assessment_code=assessment.assessment_code,
            device_code=record.device_code,
            model_code=record.model_code,
            model_version=record.model_version,
            observed_at=assessment.observed_at,
            input_uri=assessment.input_uri,
            input_size_bytes=assessment.input_size_bytes,
            input_sha256=assessment.input_sha256,
            input_summary=assessment.input_summary or {},
            target_name=assessment.target_name,
            prediction_label=assessment.prediction_label,
            confidence=float(assessment.confidence),
            prediction_basis=assessment.prediction_basis,
            status=assessment.status,
            submitted_by=assessment.submitted_by,
            submitted_by_code=assessment.submitted_by_code,
            submitted_by_role=assessment.submitted_by_role,
            review_comment=assessment.review_comment,
            reviewed_by=assessment.reviewed_by,
            reviewed_by_code=assessment.reviewed_by_code,
            reviewed_by_role=assessment.reviewed_by_role,
            reviewed_at=assessment.reviewed_at,
            created_at=assessment.created_at,
        )

    @staticmethod
    def _alert_response(record: AlertRecord) -> AlertResponse:
        """转换病虫害告警响应。

        Args:
            record: 带识别编号的告警记录。

        Returns:
            AlertResponse: 告警响应。
        """
        alert = record.alert
        return AlertResponse(
            alert_code=alert.alert_code,
            assessment_code=record.assessment_code,
            risk_level=alert.risk_level,
            message=alert.message,
            channels=alert.channels or [],
            recipients=alert.recipients or [],
            status=alert.status,
            created_by=alert.created_by,
            created_by_code=alert.created_by_code,
            created_by_role=alert.created_by_role,
            delivery_receipt_uri=alert.delivery_receipt_uri,
            delivery_receipt_size_bytes=alert.delivery_receipt_size_bytes,
            delivery_receipt_sha256=alert.delivery_receipt_sha256,
            delivered_by=alert.delivered_by,
            delivered_by_code=alert.delivered_by_code,
            delivered_by_role=alert.delivered_by_role,
            delivered_at=alert.delivered_at,
            created_at=alert.created_at,
        )

    @staticmethod
    def _consultation_response(
        record: ConsultationRecord,
    ) -> ExpertConsultationResponse:
        """转换专家会商响应。

        Args:
            record: 带报告编号的会商记录。

        Returns:
            ExpertConsultationResponse: 会商响应。
        """
        item = record.consultation
        return ExpertConsultationResponse(
            report_code=record.report_code,
            consultation_code=item.consultation_code,
            question=item.question,
            status=item.status,
            requested_by=item.requested_by,
            requested_by_code=item.requested_by_code,
            requested_by_role=item.requested_by_role,
            requested_at=item.requested_at,
            expert_organization=item.expert_organization,
            expert_title=item.expert_title,
            response=item.response,
            evidence_uri=item.evidence_uri,
            evidence_filename=item.evidence_filename,
            evidence_size_bytes=item.evidence_size_bytes,
            evidence_sha256=item.evidence_sha256,
            answered_by=item.answered_by,
            answered_by_code=item.answered_by_code,
            answered_by_role=item.answered_by_role,
            answered_at=item.answered_at,
        )

    async def _report_response(
        self,
        db: AsyncSession,
        report: PestReport,
    ) -> PestReportResponse:
        """查询报告台账和会商后转换完整响应。

        Args:
            db: 异步数据库会话。
            report: 报告模型。

        Returns:
            PestReportResponse: 完整报告响应。
        """
        items = list(await self.dao.list_report_items(db, report.id))
        consultations = await self.dao.list_consultation_records(
            db,
            report.project_id,
            report.id,
        )
        return PestReportResponse(
            report_code=report.report_code,
            report_title=report.report_title,
            scope_level=report.scope_level,
            region_code=report.region_code,
            region_name=report.region_name,
            period_start=report.period_start,
            period_end=report.period_end,
            summary=report.summary,
            conclusion=report.conclusion,
            status=report.status,
            revision_number=report.revision_number,
            assessment_count=report.assessment_count,
            alert_count=report.alert_count,
            snapshot_at=report.snapshot_at,
            file_uri=report.file_uri,
            original_filename=report.original_filename,
            file_size_bytes=report.file_size_bytes,
            checksum_sha256=report.checksum_sha256,
            created_by=report.created_by,
            created_by_code=report.created_by_code,
            created_by_role=report.created_by_role,
            last_review_comment=report.last_review_comment,
            approved_by=report.approved_by,
            approved_by_code=report.approved_by_code,
            approved_by_role=report.approved_by_role,
            approved_at=report.approved_at,
            created_at=report.created_at,
            updated_at=report.updated_at,
            items=[
                PestReportItemResponse(
                    assessment_code=item.assessment_code,
                    district_code=item.district_code,
                    district_name=item.district_name,
                    snapshot=item.snapshot or {},
                )
                for item in items
            ],
            consultation_count=len(consultations),
            open_consultation_count=sum(
                item.consultation.status == "open" for item in consultations
            ),
            download_url=(
                f"/api/v1/monitoring-network/reports/{report.report_code}/download"
                if report.status == "approved" and report.file_uri
                else None
            ),
        )

    async def _prepare_report_records(
        self,
        db: AsyncSession,
        project_id: int,
        scope_level: str,
        region_code: str,
        period_start: date,
        period_end: date,
        assessment_codes: list[str],
    ) -> tuple[str, list[ReportAssessmentRecord], int, datetime]:
        """校验报告行政范围、周期和人工批准识别结果。

        Args:
            db: 异步数据库会话。
            project_id: 项目主键。
            scope_level: 报告行政层级。
            region_code: 行政区编码。
            period_start: 报告开始日期。
            period_end: 报告结束日期。
            assessment_codes: 显式识别编号。

        Returns:
            tuple[str, list[ReportAssessmentRecord], int, datetime]: 区域名称、
                已验证识别记录、告警数和快照时间。
        """
        boundary_level = {
            "province": "province",
            "prefecture": "city",
            "county": "district",
        }[scope_level]
        boundary = await self.dao.get_region_boundary(
            db,
            project_id,
            region_code,
            boundary_level,
        )
        if boundary is None:
            raise ValidationException("报告行政区编码与申报层级不匹配")
        records = await self.dao.get_report_assessment_records(
            db,
            project_id,
            assessment_codes,
        )
        record_map = {
            item.assessment.assessment_code: item
            for item in records
        }
        missing_codes = [code for code in assessment_codes if code not in record_map]
        if missing_codes:
            raise ValidationException(f"未找到识别结果 {missing_codes[0]}")
        ordered = [record_map[code] for code in assessment_codes]
        for record in ordered:
            assessment = record.assessment
            station = record.station
            if assessment.status != "approved":
                raise ValidationException(
                    f"识别结果 {assessment.assessment_code} 尚未人工批准"
                )
            if station is None:
                raise ValidationException(
                    f"识别结果 {assessment.assessment_code} 未关联具有行政位置的设备"
                )
            observed_date = assessment.observed_at.date()
            if not period_start <= observed_date <= period_end:
                raise ValidationException(
                    f"识别结果 {assessment.assessment_code} 不在报告周期内"
                )
            matched_code = {
                "province": station.province_code,
                "prefecture": station.city_code,
                "county": station.district_code,
            }[scope_level]
            if matched_code != region_code:
                raise ValidationException(
                    f"识别结果 {assessment.assessment_code} 不属于报告行政范围"
                )
        return (
            boundary.boundary_name,
            ordered,
            sum(item.alert is not None for item in ordered),
            datetime.now(UTC),
        )

    @staticmethod
    def _report_items(
        report_id: int,
        records: Sequence[ReportAssessmentRecord],
    ) -> list[PestReportItem]:
        """把已验证识别记录固化为报告台账快照。

        Args:
            report_id: 报告主键。
            records: 已验证识别记录。

        Returns:
            list[PestReportItem]: 不可变台账条目。
        """
        return [
            PestReportItem(
                report_id=report_id,
                assessment_id=record.assessment.id,
                assessment_code=record.assessment.assessment_code,
                district_code=record.station.district_code,
                district_name=record.station.district_name,
                snapshot={
                    "observed_at": record.assessment.observed_at.isoformat(),
                    "station_code": record.station.station_code,
                    "station_name": record.station.station_name,
                    "city_code": record.station.city_code,
                    "city_name": record.station.city_name,
                    "district_code": record.station.district_code,
                    "district_name": record.station.district_name,
                    "model_code": record.model_code,
                    "model_version": record.model_version,
                    "target_name": record.assessment.target_name,
                    "prediction_label": record.assessment.prediction_label,
                    "confidence": float(record.assessment.confidence),
                    "prediction_basis": record.assessment.prediction_basis,
                    "input_uri": record.assessment.input_uri,
                    "input_sha256": record.assessment.input_sha256,
                    "reviewed_by_code": record.assessment.reviewed_by_code,
                    "alert_code": record.alert.alert_code if record.alert else None,
                    "risk_level": record.alert.risk_level if record.alert else None,
                    "alert_status": record.alert.status if record.alert else None,
                },
            )
            for record in records
            if record.station is not None
        ]

    @staticmethod
    def _write_atomic(path: Path, content: bytes) -> None:
        """通过临时文件原子写入报告实体。

        Args:
            path: 最终路径。
            content: 文件字节。

        Returns:
            None: 无返回值。
        """
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        try:
            temporary.write_bytes(content)
            os.replace(temporary, path)
        finally:
            temporary.unlink(missing_ok=True)

    def _store_evidence_sync(
        self,
        consultation_code: str,
        original_filename: str,
        file_object: BinaryIO,
    ) -> StoredEvidence:
        """流式保存专家会商实体并计算服务端校验值。

        Args:
            consultation_code: 会商编号。
            original_filename: 原始文件名。
            file_object: 上传文件流。

        Returns:
            StoredEvidence: 受控实体证据。
        """
        safe_name = Path(original_filename).name
        suffix = Path(safe_name).suffix.lower()
        if suffix not in {".pdf", ".docx", ".xlsx", ".jpg", ".jpeg", ".png", ".zip"}:
            raise ValidationException("会商证据仅支持 PDF、Office、图片或 ZIP")
        digest = hashlib.sha256()
        content = io.BytesIO()
        file_size = 0
        while chunk := file_object.read(1024 * 1024):
            file_size += len(chunk)
            if file_size > settings.max_consultation_evidence_bytes:
                raise ValidationException("会商证据超过允许大小")
            digest.update(chunk)
            content.write(chunk)
        if file_size <= 0:
            raise ValidationException("会商证据文件不能为空")
        checksum = digest.hexdigest()
        relative_path = (
            Path("consultations") / consultation_code / f"{checksum}{suffix}"
        )
        final_path = (self.storage_root / relative_path).resolve()
        if self.storage_root not in final_path.parents:
            raise ValidationException("会商证据路径越界")
        created_new = not final_path.exists()
        if created_new:
            self._write_atomic(final_path, content.getvalue())
        elif final_path.stat().st_size != file_size:
            raise ValidationException("既有会商证据大小与校验值冲突")
        return StoredEvidence(
            path=final_path,
            uri=f"storage://monitoring-reports/{relative_path.as_posix()}",
            filename=safe_name,
            file_size_bytes=file_size,
            checksum_sha256=checksum,
            created_new=created_new,
        )

    async def get_overview(
        self,
        db: AsyncSession,
        project_code: str,
        operator_code: str,
    ) -> MonitoringOverviewResponse:
        """查询田间监测网络与病虫害预警真实总览。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            operator_code: 当前项目用户编码。

        Returns:
            MonitoringOverviewResponse: 监测业务总览。
        """
        project = await self._require_project(db, project_code)
        await self.user_service.require_capability(
            db,
            project.id,
            operator_code,
            "view_monitoring_network",
        )
        stations = list(await self.dao.list_stations(db, project.id))
        devices = await self.dao.list_device_records(db, project.id)
        telemetry = await self.dao.list_telemetry_records(db, project.id)
        faults = await self.dao.list_fault_records(db, project.id)
        models = list(await self.dao.list_model_versions(db, project.id))
        assessments = await self.dao.list_assessment_records(db, project.id)
        alerts = await self.dao.list_alert_records(db, project.id)
        reports = list(await self.dao.list_reports(db, project.id))
        consultations = await self.dao.list_consultation_records(db, project.id)
        events = list(await self.dao.list_events(db, project.id))
        telemetry_count = await self.dao.count_telemetry(db, project.id)
        return MonitoringOverviewResponse(
            station_count=len(stations),
            device_count=len(devices),
            online_device_count=sum(
                record.device.status == "online" for record in devices
            ),
            abnormal_device_count=sum(
                record.device.status == "abnormal" for record in devices
            ),
            telemetry_count=telemetry_count,
            open_fault_count=sum(record.fault.status == "open" for record in faults),
            active_model_count=sum(model.status == "active" for model in models),
            pending_assessment_count=sum(
                record.assessment.status == "pending_review"
                for record in assessments
            ),
            pending_alert_count=sum(
                record.alert.status == "pending" for record in alerts
            ),
            report_count=len(reports),
            pending_report_count=sum(
                report.status not in {"approved", "draft"} for report in reports
            ),
            open_consultation_count=sum(
                record.consultation.status == "open" for record in consultations
            ),
            stations=[self._station_response(item) for item in stations],
            devices=[self._device_response(item) for item in devices],
            telemetry=[self._telemetry_response(item) for item in telemetry],
            faults=[self._fault_response(item) for item in faults],
            models=[self._model_response(item) for item in models],
            assessments=[self._assessment_response(item) for item in assessments],
            alerts=[self._alert_response(item) for item in alerts],
            reports=[
                await self._report_response(db, report)
                for report in reports
            ],
            consultations=[
                self._consultation_response(item) for item in consultations
            ],
            events=[
                MonitoringEventResponse(
                    entity_type=item.entity_type,
                    entity_code=item.entity_code,
                    event_type=item.event_type,
                    detail=item.detail or {},
                    actor=item.actor,
                    actor_code=item.actor_code,
                    actor_role=item.actor_role,
                    created_at=item.created_at,
                )
                for item in events
            ],
        )

    async def create_station(
        self,
        db: AsyncSession,
        project_code: str,
        request: StationCreateRequest,
    ) -> StationResponse:
        """登记坐标落在真实县区内的监测站。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            request: 站点来源和实体证据。

        Returns:
            StationResponse: 已登记站点。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_devices",
        )
        if await self.dao.get_station_by_code(db, project.id, request.station_code):
            raise ValidationException("监测站编号已存在")
        context = await self.dao.get_administrative_context(
            db,
            project.id,
            request.district_code,
            request.longitude,
            request.latitude,
        )
        if context is None:
            raise ValidationException("站点坐标不在申报县区真实边界内")
        station = await self.dao.add_station(
            db,
            MonitoringStation(
                project_id=project.id,
                station_code=request.station_code,
                station_name=request.station_name,
                province_code=context.province_code,
                province_name=context.province_name,
                city_code=context.city_code,
                city_name=context.city_name,
                district_code=context.district_code,
                district_name=context.district_name,
                longitude=Decimal(str(request.longitude)),
                latitude=Decimal(str(request.latitude)),
                station_type=request.station_type,
                owner_department=request.owner_department,
                source_name=request.source_name,
                source_uri=request.source_uri,
                source_version=request.source_version,
                evidence_uri=request.evidence_uri,
                evidence_size_bytes=request.evidence_size_bytes,
                evidence_sha256=request.evidence_sha256,
                status=request.status,
                registered_by=user.display_name,
                registered_by_code=user.user_code,
                registered_by_role=user.role_code,
            ),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "station",
                station.station_code,
                "station_registered",
                {
                    "district_code": station.district_code,
                    "longitude": float(station.longitude),
                    "latitude": float(station.latitude),
                    "evidence_sha256": station.evidence_sha256,
                },
                user,
            ),
        )
        await db.commit()
        return self._station_response(station)

    async def create_device(
        self,
        db: AsyncSession,
        project_code: str,
        station_code: str,
        request: DeviceCreateRequest,
    ) -> DeviceResponse:
        """在活动监测站下登记真实设备。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            station_code: 监测站编号。
            request: 设备身份、照片和操作人。

        Returns:
            DeviceResponse: 已登记设备。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_devices",
        )
        station = await self.dao.get_station_by_code(db, project.id, station_code)
        if station is None:
            raise NotFoundException("未找到监测站")
        if station.status == "retired":
            raise ValidationException("已退役监测站不能登记新设备")
        if await self.dao.get_device_by_code(db, project.id, request.device_code):
            raise ValidationException("设备编号已存在")
        device = await self.dao.add_device(
            db,
            MonitoringDevice(
                project_id=project.id,
                station_id=station.id,
                device_code=request.device_code,
                device_name=request.device_name,
                device_type=request.device_type,
                vendor=request.vendor,
                model_number=request.model_number,
                serial_number=request.serial_number,
                owner_department=request.owner_department,
                installed_at=request.installed_at,
                photo_uri=request.photo_uri,
                photo_size_bytes=request.photo_size_bytes,
                photo_sha256=request.photo_sha256,
                status=request.status,
                registered_by=user.display_name,
                registered_by_code=user.user_code,
                registered_by_role=user.role_code,
            ),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "device",
                device.device_code,
                "device_registered",
                {
                    "station_code": station.station_code,
                    "device_type": device.device_type,
                    "photo_sha256": device.photo_sha256,
                },
                user,
            ),
        )
        await db.commit()
        return self._device_response(
            DeviceRecord(device=device, station_code=station.station_code)
        )

    async def ingest_telemetry(
        self,
        db: AsyncSession,
        project_code: str,
        device_code: str,
        request: TelemetryCreateRequest,
    ) -> TelemetryResponse:
        """幂等写入设备遥测并刷新在线状态。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            device_code: 设备编号。
            request: 遥测载荷、证据和幂等键。

        Returns:
            TelemetryResponse: 新增或已存在的同一遥测。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "ingest_monitoring_data",
        )
        device = await self.dao.get_device_by_code(db, project.id, device_code)
        if device is None:
            raise NotFoundException("未找到监测设备")
        if device.status == "retired":
            raise ValidationException("已退役设备不能继续写入遥测")
        canonical_payload = request.model_dump(mode="json", exclude={"operator_code"})
        request_sha256 = hashlib.sha256(
            json.dumps(
                canonical_payload,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest()
        existing = await self.dao.get_telemetry_by_idempotency_key(
            db,
            device.id,
            request.idempotency_key,
        )
        if existing is not None:
            if existing.request_sha256 != request_sha256:
                raise ValidationException("幂等键已被不同遥测载荷占用")
            return self._telemetry_response(
                TelemetryRecord(telemetry=existing, device_code=device.device_code)
            )
        telemetry = await self.dao.add_telemetry(
            db,
            DeviceTelemetry(
                device_id=device.id,
                idempotency_key=request.idempotency_key,
                request_sha256=request_sha256,
                observed_at=request.observed_at,
                metric_code=request.metric_code,
                metric_value=(
                    Decimal(str(request.metric_value))
                    if request.metric_value is not None
                    else None
                ),
                metric_unit=request.metric_unit,
                payload=request.payload,
                evidence_uri=request.evidence_uri,
                evidence_size_bytes=request.evidence_size_bytes,
                evidence_sha256=request.evidence_sha256,
                ingested_by=user.display_name,
                ingested_by_code=user.user_code,
                ingested_by_role=user.role_code,
            ),
        )
        device.last_telemetry_at = request.observed_at
        if device.status not in {"maintenance", "abnormal"}:
            device.status = "online"
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "device",
                device.device_code,
                "telemetry_ingested",
                {
                    "idempotency_key": telemetry.idempotency_key,
                    "metric_code": telemetry.metric_code,
                    "observed_at": telemetry.observed_at.isoformat(),
                },
                user,
            ),
        )
        await db.commit()
        return self._telemetry_response(
            TelemetryRecord(telemetry=telemetry, device_code=device.device_code)
        )

    async def create_fault(
        self,
        db: AsyncSession,
        project_code: str,
        device_code: str,
        request: FaultCreateRequest,
    ) -> FaultResponse:
        """登记设备故障并进入处置队列。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            device_code: 设备编号。
            request: 故障严重度、原因和发生时间。

        Returns:
            FaultResponse: 已登记故障。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "report_device_fault",
        )
        device = await self.dao.get_device_by_code(db, project.id, device_code)
        if device is None:
            raise NotFoundException("未找到监测设备")
        if device.status == "retired":
            raise ValidationException("已退役设备不能新建故障")
        if await self.dao.get_fault_by_code(db, project.id, request.fault_code):
            raise ValidationException("故障编号已存在")
        fault = await self.dao.add_fault(
            db,
            DeviceFault(
                project_id=project.id,
                device_id=device.id,
                fault_code=request.fault_code,
                severity=request.severity,
                reason=request.reason,
                occurred_at=request.occurred_at,
                status="open",
                reported_by=user.display_name,
                reported_by_code=user.user_code,
                reported_by_role=user.role_code,
            ),
        )
        device.status = "abnormal"
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "fault",
                fault.fault_code,
                "fault_reported",
                {
                    "device_code": device.device_code,
                    "severity": fault.severity,
                },
                user,
            ),
        )
        await db.commit()
        return self._fault_response(
            FaultRecord(fault=fault, device_code=device.device_code)
        )

    async def resolve_fault(
        self,
        db: AsyncSession,
        project_code: str,
        fault_code: str,
        request: FaultResolveRequest,
    ) -> FaultResponse:
        """关闭故障并保存实体处置证据。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            fault_code: 故障编号。
            request: 处置说明、实体证据和操作人。

        Returns:
            FaultResponse: 已关闭故障。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_devices",
        )
        fault = await self.dao.get_fault_by_code(
            db,
            project.id,
            fault_code,
            for_update=True,
        )
        if fault is None:
            raise NotFoundException("未找到设备故障")
        if fault.status != "open":
            raise ValidationException("设备故障已经关闭")
        device_records = await self.dao.list_device_records(db, project.id)
        record = next(
            (item for item in device_records if item.device.id == fault.device_id),
            None,
        )
        if record is None:
            raise NotFoundException("未找到故障关联设备")
        fault.status = "resolved"
        fault.resolution_comment = request.resolution_comment
        fault.resolution_evidence_uri = request.resolution_evidence_uri
        fault.resolution_evidence_size_bytes = request.resolution_evidence_size_bytes
        fault.resolution_evidence_sha256 = request.resolution_evidence_sha256
        fault.resolved_by = user.display_name
        fault.resolved_by_code = user.user_code
        fault.resolved_by_role = user.role_code
        fault.resolved_at = datetime.now(UTC)
        await db.flush()
        if await self.dao.count_open_faults_for_device(db, record.device.id) == 0:
            if record.device.status not in {"maintenance", "retired"}:
                record.device.status = (
                    "online" if record.device.last_telemetry_at else "offline"
                )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "fault",
                fault.fault_code,
                "fault_resolved",
                {
                    "device_code": record.device.device_code,
                    "resolution_evidence_sha256": request.resolution_evidence_sha256,
                },
                user,
            ),
        )
        await db.commit()
        return self._fault_response(
            FaultRecord(fault=fault, device_code=record.device.device_code)
        )

    async def register_model(
        self,
        db: AsyncSession,
        project_code: str,
        request: PestModelCreateRequest,
    ) -> PestModelResponse:
        """登记模型实体并替代同编码旧活动版本。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            request: 模型来源、实体和评估指标。

        Returns:
            PestModelResponse: 已登记模型版本。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_monitoring_models",
        )
        if await self.dao.get_model_version(
            db,
            project.id,
            request.model_code,
            request.model_version,
        ):
            raise ValidationException("模型编码和版本已存在")
        old_versions = await self.dao.list_active_model_versions_by_code(
            db,
            project.id,
            request.model_code,
        )
        for old_version in old_versions:
            old_version.status = "superseded"
            old_version.superseded_by_version = request.model_version
        model = await self.dao.add_model_version(
            db,
            PestModelVersion(
                project_id=project.id,
                model_code=request.model_code,
                model_version=request.model_version,
                model_name=request.model_name,
                target_type=request.target_type,
                deployment_target=request.deployment_target,
                training_source_uri=request.training_source_uri,
                evaluation_source_uri=request.evaluation_source_uri,
                artifact_uri=request.artifact_uri,
                artifact_size_bytes=request.artifact_size_bytes,
                artifact_sha256=request.artifact_sha256,
                accuracy=Decimal(str(request.accuracy)),
                recall=Decimal(str(request.recall)),
                f1_score=Decimal(str(request.f1_score)),
                roc_auc=Decimal(str(request.roc_auc)),
                status="active",
                registered_by=user.display_name,
                registered_by_code=user.user_code,
                registered_by_role=user.role_code,
            ),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "model",
                f"{model.model_code}:{model.model_version}",
                "model_registered",
                {
                    "artifact_sha256": model.artifact_sha256,
                    "superseded_versions": [
                        item.model_version for item in old_versions
                    ],
                    "metrics": {
                        "accuracy": float(model.accuracy),
                        "recall": float(model.recall),
                        "f1_score": float(model.f1_score),
                        "roc_auc": float(model.roc_auc),
                    },
                },
                user,
            ),
        )
        await db.commit()
        return self._model_response(model)

    async def create_assessment(
        self,
        db: AsyncSession,
        project_code: str,
        request: AssessmentCreateRequest,
    ) -> AssessmentResponse:
        """登记模型识别结果并进入人工复核队列。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            request: 模型、输入实体、预测和操作人。

        Returns:
            AssessmentResponse: 待复核识别结果。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "ingest_monitoring_data",
        )
        if await self.dao.get_assessment_by_code(
            db,
            project.id,
            request.assessment_code,
        ):
            raise ValidationException("识别结果编号已存在")
        model = await self.dao.get_model_version(
            db,
            project.id,
            request.model_code,
            request.model_version,
        )
        if model is None:
            raise NotFoundException("未找到指定模型版本")
        if model.status != "active":
            raise ValidationException("已被替代的模型版本不能新增识别结果")
        device = None
        if request.device_code:
            device = await self.dao.get_device_by_code(
                db,
                project.id,
                request.device_code,
            )
            if device is None:
                raise NotFoundException("未找到识别结果关联设备")
        assessment = await self.dao.add_assessment(
            db,
            PestAssessment(
                project_id=project.id,
                device_id=device.id if device else None,
                model_version_id=model.id,
                assessment_code=request.assessment_code,
                observed_at=request.observed_at,
                input_uri=request.input_uri,
                input_size_bytes=request.input_size_bytes,
                input_sha256=request.input_sha256,
                input_summary=request.input_summary,
                target_name=request.target_name,
                prediction_label=request.prediction_label,
                confidence=Decimal(str(request.confidence)),
                prediction_basis=request.prediction_basis,
                status="pending_review",
                submitted_by=user.display_name,
                submitted_by_code=user.user_code,
                submitted_by_role=user.role_code,
            ),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "assessment",
                assessment.assessment_code,
                "assessment_submitted",
                {
                    "model_code": model.model_code,
                    "model_version": model.model_version,
                    "input_sha256": assessment.input_sha256,
                    "confidence": float(assessment.confidence),
                },
                user,
            ),
        )
        await db.commit()
        return self._assessment_response(
            AssessmentRecord(
                assessment=assessment,
                device_code=device.device_code if device else None,
                model_code=model.model_code,
                model_version=model.model_version,
            )
        )

    async def review_assessment(
        self,
        db: AsyncSession,
        project_code: str,
        assessment_code: str,
        request: AssessmentReviewRequest,
    ) -> AssessmentResponse:
        """人工批准或驳回模型识别结果。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            assessment_code: 识别结果编号。
            request: 复核结论、依据和操作人。

        Returns:
            AssessmentResponse: 已复核识别结果。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "review_model_output",
        )
        assessment = await self.dao.get_assessment_by_code(
            db,
            project.id,
            assessment_code,
            for_update=True,
        )
        if assessment is None:
            raise NotFoundException("未找到病虫害识别结果")
        if assessment.status != "pending_review":
            raise ValidationException("识别结果已经完成复核")
        assessment.status = "approved" if request.decision == "approve" else "rejected"
        assessment.review_comment = request.comment
        assessment.reviewed_by = user.display_name
        assessment.reviewed_by_code = user.user_code
        assessment.reviewed_by_role = user.role_code
        assessment.reviewed_at = datetime.now(UTC)
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "assessment",
                assessment.assessment_code,
                "assessment_reviewed",
                {
                    "decision": request.decision,
                    "comment": request.comment,
                },
                user,
            ),
        )
        await db.commit()
        records = await self.dao.list_assessment_records(db, project.id)
        record = next(
            item
            for item in records
            if item.assessment.assessment_code == assessment.assessment_code
        )
        return self._assessment_response(record)

    async def create_alert(
        self,
        db: AsyncSession,
        project_code: str,
        assessment_code: str,
        request: AlertCreateRequest,
    ) -> AlertResponse:
        """基于已批准识别结果创建待发送告警。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            assessment_code: 已批准识别结果编号。
            request: 风险、渠道、接收对象和操作人。

        Returns:
            AlertResponse: 待送达告警。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "deliver_pest_alert",
        )
        assessment = await self.dao.get_assessment_by_code(
            db,
            project.id,
            assessment_code,
        )
        if assessment is None:
            raise NotFoundException("未找到病虫害识别结果")
        if assessment.status != "approved":
            raise ValidationException("只有人工批准的识别结果才能创建告警")
        if await self.dao.get_alert_by_assessment_id(db, assessment.id):
            raise ValidationException("该识别结果已经创建告警")
        if await self.dao.get_alert_by_code(db, project.id, request.alert_code):
            raise ValidationException("告警编号已存在")
        alert = await self.dao.add_alert(
            db,
            PestAlert(
                project_id=project.id,
                assessment_id=assessment.id,
                alert_code=request.alert_code,
                risk_level=request.risk_level,
                message=request.message,
                channels=request.channels,
                recipients=request.recipients,
                status="pending",
                created_by=user.display_name,
                created_by_code=user.user_code,
                created_by_role=user.role_code,
            ),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "alert",
                alert.alert_code,
                "alert_created",
                {
                    "assessment_code": assessment.assessment_code,
                    "risk_level": alert.risk_level,
                    "channels": alert.channels,
                },
                user,
            ),
        )
        await db.commit()
        return self._alert_response(
            AlertRecord(alert=alert, assessment_code=assessment.assessment_code)
        )

    async def deliver_alert(
        self,
        db: AsyncSession,
        project_code: str,
        alert_code: str,
        request: AlertDeliverRequest,
    ) -> AlertResponse:
        """登记告警真实送达回执并完成闭环。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            alert_code: 告警编号。
            request: 回执实体和操作人。

        Returns:
            AlertResponse: 已送达告警。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "deliver_pest_alert",
        )
        alert = await self.dao.get_alert_by_code(
            db,
            project.id,
            alert_code,
            for_update=True,
        )
        if alert is None:
            raise NotFoundException("未找到病虫害告警")
        if alert.status != "pending":
            raise ValidationException("告警已经登记送达回执")
        alert.status = "delivered"
        alert.delivery_receipt_uri = request.delivery_receipt_uri
        alert.delivery_receipt_size_bytes = request.delivery_receipt_size_bytes
        alert.delivery_receipt_sha256 = request.delivery_receipt_sha256
        alert.delivered_by = user.display_name
        alert.delivered_by_code = user.user_code
        alert.delivered_by_role = user.role_code
        alert.delivered_at = datetime.now(UTC)
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "alert",
                alert.alert_code,
                "alert_delivered",
                {
                    "delivery_receipt_sha256": request.delivery_receipt_sha256,
                    "channels": alert.channels,
                    "recipients": alert.recipients,
                },
                user,
            ),
        )
        await db.commit()
        records = await self.dao.list_alert_records(db, project.id)
        record = next(
            item for item in records if item.alert.alert_code == alert.alert_code
        )
        return self._alert_response(record)

    async def create_report(
        self,
        db: AsyncSession,
        project_code: str,
        request: PestReportCreateRequest,
    ) -> PestReportResponse:
        """从已人工批准识别结果创建病虫害报告草稿。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            request: 报告范围、周期、内容和识别编号。

        Returns:
            PestReportResponse: 已创建报告草稿。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_pest_reports",
        )
        if await self.dao.get_report_by_code(db, project.id, request.report_code):
            raise ValidationException("报告编号已存在")
        region_name, records, alert_count, snapshot_at = (
            await self._prepare_report_records(
                db,
                project.id,
                request.scope_level,
                request.region_code,
                request.period_start,
                request.period_end,
                request.assessment_codes,
            )
        )
        report = await self.dao.add_report(
            db,
            PestReport(
                project_id=project.id,
                report_code=request.report_code,
                report_title=request.report_title.strip(),
                scope_level=request.scope_level,
                region_code=request.region_code.strip(),
                region_name=region_name,
                period_start=request.period_start,
                period_end=request.period_end,
                summary=request.summary.strip(),
                conclusion=request.conclusion.strip(),
                status="draft",
                revision_number=1,
                assessment_count=len(records),
                alert_count=alert_count,
                snapshot_at=snapshot_at,
                created_by=user.display_name,
                created_by_code=user.user_code,
                created_by_role=user.role_code,
            ),
        )
        await self.dao.replace_report_items(
            db,
            report.id,
            self._report_items(report.id, records),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "pest_report",
                report.report_code,
                "report_created",
                {
                    "scope_level": report.scope_level,
                    "region_code": report.region_code,
                    "assessment_count": report.assessment_count,
                    "alert_count": report.alert_count,
                    "revision_number": report.revision_number,
                },
                user,
            ),
        )
        await db.commit()
        return await self._report_response(db, report)

    async def revise_report(
        self,
        db: AsyncSession,
        project_code: str,
        report_code: str,
        request: PestReportReviseRequest,
    ) -> PestReportResponse:
        """修订草稿或退回后的报告并重建显式台账快照。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            report_code: 报告编号。
            request: 修订内容、依据和操作人。

        Returns:
            PestReportResponse: 修订后报告。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_pest_reports",
        )
        report = await self.dao.get_report_by_code(
            db,
            project.id,
            report_code,
            for_update=True,
        )
        if report is None:
            raise NotFoundException("未找到病虫害报告")
        if report.status not in {"draft", "returned"}:
            raise ValidationException("只有草稿或退回报告可以修订")
        region_name, records, alert_count, snapshot_at = (
            await self._prepare_report_records(
                db,
                project.id,
                request.scope_level,
                request.region_code,
                request.period_start,
                request.period_end,
                request.assessment_codes,
            )
        )
        previous_status = report.status
        report.report_title = request.report_title.strip()
        report.scope_level = request.scope_level
        report.region_code = request.region_code.strip()
        report.region_name = region_name
        report.period_start = request.period_start
        report.period_end = request.period_end
        report.summary = request.summary.strip()
        report.conclusion = request.conclusion.strip()
        report.status = "draft"
        report.revision_number += 1
        report.assessment_count = len(records)
        report.alert_count = alert_count
        report.snapshot_at = snapshot_at
        report.file_uri = None
        report.original_filename = None
        report.file_size_bytes = None
        report.checksum_sha256 = None
        report.approved_by = None
        report.approved_by_code = None
        report.approved_by_role = None
        report.approved_at = None
        report.last_review_comment = request.revision_comment.strip()
        await self.dao.replace_report_items(
            db,
            report.id,
            self._report_items(report.id, records),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "pest_report",
                report.report_code,
                "report_revised",
                {
                    "previous_status": previous_status,
                    "revision_number": report.revision_number,
                    "assessment_count": report.assessment_count,
                    "comment": request.revision_comment.strip(),
                },
                user,
            ),
        )
        await db.commit()
        return await self._report_response(db, report)

    async def create_consultation(
        self,
        db: AsyncSession,
        project_code: str,
        report_code: str,
        request: ExpertConsultationCreateRequest,
    ) -> ExpertConsultationResponse:
        """为草稿或退回报告发起专家会商。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            report_code: 报告编号。
            request: 会商编号、问题和操作人。

        Returns:
            ExpertConsultationResponse: 待答复会商。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_pest_reports",
        )
        report = await self.dao.get_report_by_code(
            db,
            project.id,
            report_code,
            for_update=True,
        )
        if report is None:
            raise NotFoundException("未找到病虫害报告")
        if report.status not in {"draft", "returned"}:
            raise ValidationException("只有草稿或退回报告可以发起会商")
        if await self.dao.get_consultation_by_code(
            db,
            project.id,
            request.consultation_code,
        ):
            raise ValidationException("会商编号已存在")
        consultation = await self.dao.add_consultation(
            db,
            ExpertConsultation(
                project_id=project.id,
                report_id=report.id,
                consultation_code=request.consultation_code,
                question=request.question.strip(),
                status="open",
                requested_by=user.display_name,
                requested_by_code=user.user_code,
                requested_by_role=user.role_code,
            ),
        )
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "pest_report",
                report.report_code,
                "expert_consultation_requested",
                {
                    "consultation_code": consultation.consultation_code,
                    "question": consultation.question,
                },
                user,
            ),
        )
        await db.commit()
        return self._consultation_response(
            ConsultationRecord(
                consultation=consultation,
                report_code=report.report_code,
            )
        )

    async def answer_consultation(
        self,
        db: AsyncSession,
        project_code: str,
        consultation_code: str,
        request: ExpertConsultationAnswerRequest,
        evidence_filename: str,
        evidence_file: BinaryIO,
    ) -> ExpertConsultationResponse:
        """上传会商答复实体并登记稳定答复人身份。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            consultation_code: 会商编号。
            request: 专家单位、职称、答复和操作人。
            evidence_filename: 实体证据原始文件名。
            evidence_file: 上传文件流。

        Returns:
            ExpertConsultationResponse: 已答复会商。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "answer_pest_consultation",
        )
        consultation = await self.dao.get_consultation_by_code(
            db,
            project.id,
            consultation_code,
            for_update=True,
        )
        if consultation is None:
            raise NotFoundException("未找到专家会商")
        if consultation.status != "open":
            raise ValidationException("专家会商已经答复")
        report = await self.dao.get_report_by_id(
            db,
            project.id,
            consultation.report_id,
        )
        if report is None or report.status not in {"draft", "returned"}:
            raise ValidationException("报告已进入审核，不能补登记会商答复")
        stored = await asyncio.to_thread(
            self._store_evidence_sync,
            consultation.consultation_code,
            evidence_filename,
            evidence_file,
        )
        committed = False
        try:
            consultation.status = "answered"
            consultation.expert_organization = request.expert_organization.strip()
            consultation.expert_title = request.expert_title.strip()
            consultation.response = request.response.strip()
            consultation.evidence_uri = stored.uri
            consultation.evidence_filename = stored.filename
            consultation.evidence_size_bytes = stored.file_size_bytes
            consultation.evidence_sha256 = stored.checksum_sha256
            consultation.answered_by = user.display_name
            consultation.answered_by_code = user.user_code
            consultation.answered_by_role = user.role_code
            consultation.answered_at = datetime.now(UTC)
            await self.dao.add_event(
                db,
                self._event(
                    project.id,
                    "pest_report",
                    report.report_code,
                    "expert_consultation_answered",
                    {
                        "consultation_code": consultation.consultation_code,
                        "evidence_sha256": stored.checksum_sha256,
                        "expert_organization": consultation.expert_organization,
                        "expert_title": consultation.expert_title,
                    },
                    user,
                ),
            )
            await db.commit()
            committed = True
        finally:
            if not committed and stored.created_new:
                stored.path.unlink(missing_ok=True)
        return self._consultation_response(
            ConsultationRecord(
                consultation=consultation,
                report_code=report.report_code,
            )
        )

    async def submit_report(
        self,
        db: AsyncSession,
        project_code: str,
        report_code: str,
        request: PestReportSubmitRequest,
    ) -> PestReportResponse:
        """提交报告并由县级审核阶段开始流转。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            report_code: 报告编号。
            request: 提交说明和操作人。

        Returns:
            PestReportResponse: 已进入县级审核的报告。
        """
        project = await self._require_project(db, project_code)
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            "manage_pest_reports",
        )
        report = await self.dao.get_report_by_code(
            db,
            project.id,
            report_code,
            for_update=True,
        )
        if report is None:
            raise NotFoundException("未找到病虫害报告")
        if report.status not in {"draft", "returned"}:
            raise ValidationException("当前报告状态不能提交")
        consultations = await self.dao.list_consultation_records(
            db,
            project.id,
            report.id,
        )
        if any(item.consultation.status == "open" for item in consultations):
            raise ValidationException("存在未答复专家会商，不能提交审核")
        if report.assessment_count <= 0:
            raise ValidationException("报告没有显式识别台账，不能提交")
        previous_status = report.status
        report.status = "county_review"
        report.last_review_comment = request.comment.strip()
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "pest_report",
                report.report_code,
                "report_submitted",
                {
                    "previous_status": previous_status,
                    "next_status": report.status,
                    "revision_number": report.revision_number,
                    "comment": request.comment.strip(),
                },
                user,
            ),
        )
        await db.commit()
        return await self._report_response(db, report)

    @staticmethod
    def _build_report_workbook(
        report: PestReport,
        items: Sequence[PestReportItem],
        consultations: Sequence[ConsultationRecord],
        events: Sequence[MonitoringEvent],
    ) -> bytes:
        """生成包含摘要、识别台账、审核和会商证据的 XLSX。

        Args:
            report: 已通过省级审核的报告。
            items: 显式识别台账。
            consultations: 专家会商记录。
            events: 报告审核事件。

        Returns:
            bytes: Excel 实体字节。
        """
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "报告摘要"
        summary_rows = [
            ("报告编号", report.report_code),
            ("报告标题", report.report_title),
            ("行政层级", report.scope_level),
            ("行政区域", f"{report.region_name}（{report.region_code}）"),
            ("监测周期", f"{report.period_start} 至 {report.period_end}"),
            ("版本", report.revision_number),
            ("识别结果数", report.assessment_count),
            ("关联告警数", report.alert_count),
            ("摘要", report.summary),
            ("结论与建议", report.conclusion),
            ("批准人", report.approved_by or ""),
            ("批准人编码", report.approved_by_code or ""),
            ("批准时间", report.approved_at.isoformat() if report.approved_at else ""),
        ]
        for row in summary_rows:
            summary_sheet.append(row)
        summary_sheet.column_dimensions["A"].width = 18
        summary_sheet.column_dimensions["B"].width = 90

        ledger = workbook.create_sheet("识别台账")
        ledger.append([
            "识别编号",
            "地级区域",
            "县区",
            "观测时间",
            "模型版本",
            "对象",
            "预测结论",
            "置信度",
            "风险等级",
            "告警状态",
            "输入 SHA-256",
            "人工复核人编码",
        ])
        for item in items:
            snapshot = item.snapshot or {}
            ledger.append([
                item.assessment_code,
                snapshot.get("city_name", ""),
                item.district_name,
                snapshot.get("observed_at", ""),
                f"{snapshot.get('model_code', '')}:{snapshot.get('model_version', '')}",
                snapshot.get("target_name", ""),
                snapshot.get("prediction_label", ""),
                snapshot.get("confidence", ""),
                snapshot.get("risk_level", ""),
                snapshot.get("alert_status", ""),
                snapshot.get("input_sha256", ""),
                snapshot.get("reviewed_by_code", ""),
            ])

        audit = workbook.create_sheet("审核与会商")
        audit.append([
            "类型",
            "编号/动作",
            "人员",
            "角色/单位",
            "时间",
            "依据/答复",
            "证据 SHA-256",
        ])
        for event in reversed(events):
            audit.append([
                "审核事件",
                event.event_type,
                event.actor,
                event.actor_role,
                event.created_at.isoformat(),
                json.dumps(event.detail or {}, ensure_ascii=False),
                "",
            ])
        for record in reversed(consultations):
            item = record.consultation
            audit.append([
                "专家会商",
                item.consultation_code,
                item.answered_by or item.requested_by,
                item.expert_organization or item.requested_by_role,
                (item.answered_at or item.requested_at).isoformat(),
                item.response or item.question,
                item.evidence_sha256 or "",
            ])
        content = io.BytesIO()
        workbook.save(content)
        return content.getvalue()

    async def review_report(
        self,
        db: AsyncSession,
        project_code: str,
        report_code: str,
        request: PestReportReviewRequest,
    ) -> PestReportResponse:
        """执行县、市、省分级审核，省级通过后生成实体报告。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            report_code: 报告编号。
            request: 审核动作、依据和稳定用户。

        Returns:
            PestReportResponse: 审核后报告。
        """
        project = await self._require_project(db, project_code)
        report = await self.dao.get_report_by_code(
            db,
            project.id,
            report_code,
            for_update=True,
        )
        if report is None:
            raise NotFoundException("未找到病虫害报告")
        stage = {
            "county_review": ("review_county_pest_report", "prefecture_review"),
            "prefecture_review": ("review_prefecture_pest_report", "province_review"),
            "province_review": ("review_province_pest_report", "approved"),
        }.get(report.status)
        if stage is None:
            raise ValidationException("当前报告状态不允许审核")
        user = await self.user_service.require_capability(
            db,
            project.id,
            request.operator_code,
            stage[0],
        )
        previous_status = report.status
        report.last_review_comment = request.comment.strip()
        report.status = "returned" if request.action == "return" else stage[1]
        if report.status == "approved":
            report.approved_by = user.display_name
            report.approved_by_code = user.user_code
            report.approved_by_role = user.role_code
            report.approved_at = datetime.now(UTC)
        await self.dao.add_event(
            db,
            self._event(
                project.id,
                "pest_report",
                report.report_code,
                (
                    "report_returned"
                    if request.action == "return"
                    else "report_approved_stage"
                ),
                {
                    "previous_status": previous_status,
                    "next_status": report.status,
                    "comment": request.comment.strip(),
                    "revision_number": report.revision_number,
                },
                user,
            ),
        )
        generated_path: Path | None = None
        if report.status == "approved":
            items = list(await self.dao.list_report_items(db, report.id))
            consultations = await self.dao.list_consultation_records(
                db,
                project.id,
                report.id,
            )
            events = [
                item
                for item in await self.dao.list_events(db, project.id, limit=500)
                if item.entity_type == "pest_report"
                and item.entity_code == report.report_code
            ]
            content = self._build_report_workbook(
                report,
                items,
                consultations,
                events,
            )
            checksum = hashlib.sha256(content).hexdigest()
            filename = f"{report.report_code}-v{report.revision_number}.xlsx"
            relative_path = Path("reports") / report.report_code / filename
            generated_path = (self.storage_root / relative_path).resolve()
            if self.storage_root not in generated_path.parents:
                raise ValidationException("报告实体路径越界")
            await asyncio.to_thread(self._write_atomic, generated_path, content)
            report.file_uri = (
                f"storage://monitoring-reports/{relative_path.as_posix()}"
            )
            report.original_filename = filename
            report.file_size_bytes = len(content)
            report.checksum_sha256 = checksum
        committed = False
        try:
            await db.commit()
            committed = True
        finally:
            if not committed and generated_path is not None:
                generated_path.unlink(missing_ok=True)
        return await self._report_response(db, report)

    def _verify_report_file(self, report: PestReport) -> Path:
        """复核报告受控路径、大小和 SHA-256。

        Args:
            report: 已批准报告。

        Returns:
            Path: 校验通过的实体路径。
        """
        prefix = "storage://monitoring-reports/"
        if not report.file_uri or not report.file_uri.startswith(prefix):
            raise ValidationException("报告实体地址不受控")
        path = (self.storage_root / report.file_uri.removeprefix(prefix)).resolve()
        if self.storage_root not in path.parents or not path.is_file():
            raise ValidationException("报告实体不存在或路径越界")
        if path.stat().st_size != report.file_size_bytes:
            raise ValidationException("报告实体大小校验失败")
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if digest != report.checksum_sha256:
            raise ValidationException("报告实体 SHA-256 校验失败")
        return path

    async def get_report_download(
        self,
        db: AsyncSession,
        project_code: str,
        report_code: str,
        operator_code: str,
    ) -> ReportDownload:
        """鉴权并复核实体后返回病虫害报告下载信息。

        Args:
            db: 异步数据库会话。
            project_code: 项目编号。
            report_code: 报告编号。
            operator_code: 当前项目用户编码。

        Returns:
            ReportDownload: 下载实体信息。
        """
        project = await self._require_project(db, project_code)
        await self.user_service.require_capability(
            db,
            project.id,
            operator_code,
            "download_pest_report",
        )
        report = await self.dao.get_report_by_code(db, project.id, report_code)
        if report is None:
            raise NotFoundException("未找到病虫害报告")
        if report.status != "approved" or not report.original_filename:
            raise ValidationException("报告尚未通过省级审核")
        path = await asyncio.to_thread(self._verify_report_file, report)
        return ReportDownload(
            path=path,
            filename=report.original_filename,
            media_type=(
                mimetypes.guess_type(report.original_filename)[0]
                or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            checksum_sha256=report.checksum_sha256 or "",
        )
