"""灾害监测专题报告生成、校验、下载与审计服务。"""

import asyncio
import os
import secrets
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundException, ValidationException
from app.core.imagery_files import calculate_sha256
from app.dao.disaster_dao import DisasterDAO
from app.dao.disaster_report_dao import DisasterReportDAO
from app.dao.workbench_dao import WorkbenchDAO
from app.models.disaster_report import DisasterReport
from app.models.workbench import ReviewRecord
from app.schemas.disaster import (
    DisasterReportGenerateRequest,
    DisasterReportListResponse,
    DisasterReportResponse,
    DisasterSummaryResponse,
)
from app.services.disaster_service import DisasterService
from app.services.project_user_service import ProjectUserService

SEVERITY_COLORS = {
    "轻度": "#F2C94C",
    "中度": "#F2994A",
    "重度": "#EB5757",
    "绝收": "#8B1E1E",
}


@dataclass(frozen=True)
class DisasterReportDownload:
    """通过实体复核的灾害专题报告下载信息。"""

    path: Path
    filename: str
    checksum_sha256: str


class DisasterReportService:
    """生成含统计、图表、分布图和来源审计的 XLSX 报告。"""

    def __init__(
        self,
        dao: DisasterReportDAO | None = None,
        disaster_dao: DisasterDAO | None = None,
        workbench_dao: WorkbenchDAO | None = None,
        project_user_service: ProjectUserService | None = None,
        disaster_service: DisasterService | None = None,
        storage_root: Path | None = None,
    ) -> None:
        """初始化灾害专题报告服务。

        Args:
            dao: 灾害报告 DAO。
            disaster_dao: 灾害斑块 DAO。
            workbench_dao: 工作台公共 DAO。
            project_user_service: 项目用户权限服务。
            disaster_service: 灾害汇总服务。
            storage_root: 可注入受控存储根目录。

        Returns:
            None: 无返回值。
        """
        self.dao = dao or DisasterReportDAO()
        self.disaster_dao = disaster_dao or DisasterDAO()
        self.workbench_dao = workbench_dao or WorkbenchDAO()
        self.project_user_service = project_user_service or ProjectUserService()
        self.disaster_service = disaster_service or DisasterService()
        self.storage_root = storage_root or (
            Path(__file__).resolve().parents[2] / "storage" / "disaster-reports"
        )

    def _resolve_path(self, relative_path: str) -> Path:
        """解析并约束灾害报告受控路径。

        Args:
            relative_path: 相对于报告根目录的路径。

        Returns:
            Path: 未越界的绝对路径。
        """
        root = self.storage_root.resolve()
        path = (root / relative_path).resolve()
        if not path.is_relative_to(root):
            raise ValidationException("灾害专题报告存储路径越界")
        return path

    @staticmethod
    def _load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        """加载可显示中文的系统字体并提供安全回退。

        Args:
            size: 字体像素大小。

        Returns:
            ImageFont.FreeTypeFont | ImageFont.ImageFont: 可用字体。
        """
        candidates = (
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
        )
        for candidate in candidates:
            if Path(candidate).is_file():
                return ImageFont.truetype(candidate, size=size)
        return ImageFont.load_default()

    @classmethod
    def _render_distribution_map(cls, summary: DisasterSummaryResponse) -> bytes:
        """把已复核灾害 Polygon 渲染为报告内嵌分布图。

        Args:
            summary: 灾害汇总及 GeoJSON 几何。

        Returns:
            bytes: PNG 实体字节。
        """
        width, height = 1400, 900
        margin = 90
        image = Image.new("RGB", (width, height), "#F6F7F5")
        draw = ImageDraw.Draw(image)
        title_font = cls._load_font(34)
        label_font = cls._load_font(22)
        small_font = cls._load_font(18)
        draw.text((margin, 25), "灾害斑块空间分布图", fill="#25342C", font=title_font)
        included = [item for item in summary.items if item.status != "excluded"]
        coordinates = [
            coordinate
            for item in included
            for ring in item.geometry.get("coordinates", [])
            for coordinate in ring
        ]
        if not coordinates:
            draw.text(
                (margin, 160),
                "复核后无纳入统计的灾害斑块",
                fill="#6B7770",
                font=label_font,
            )
        else:
            min_lon = min(float(point[0]) for point in coordinates)
            max_lon = max(float(point[0]) for point in coordinates)
            min_lat = min(float(point[1]) for point in coordinates)
            max_lat = max(float(point[1]) for point in coordinates)
            lon_span = max(max_lon - min_lon, 1e-8)
            lat_span = max(max_lat - min_lat, 1e-8)
            map_top = 100
            map_bottom = height - 90
            map_width = width - margin * 2
            map_height = map_bottom - map_top

            def project(point: list[float] | tuple[float, float]) -> tuple[int, int]:
                x = margin + (float(point[0]) - min_lon) / lon_span * map_width
                y = map_bottom - (float(point[1]) - min_lat) / lat_span * map_height
                return round(x), round(y)

            for item in included:
                color = SEVERITY_COLORS.get(item.severity, "#8A9690")
                for ring_index, ring in enumerate(
                    item.geometry.get("coordinates", [])
                ):
                    projected = [project(point) for point in ring]
                    if len(projected) < 3:
                        continue
                    if ring_index == 0:
                        draw.polygon(projected, fill=color, outline="#5E332C", width=2)
                    else:
                        draw.polygon(
                            projected,
                            fill="#F6F7F5",
                            outline="#5E332C",
                            width=1,
                        )
            draw.rectangle(
                (margin, map_top, width - margin, map_bottom),
                outline="#839087",
                width=2,
            )
        legend_x = width - 325
        legend_y = 34
        for index, (severity, color) in enumerate(SEVERITY_COLORS.items()):
            top = legend_y + index * 30
            draw.rectangle((legend_x, top, legend_x + 22, top + 18), fill=color)
            draw.text(
                (legend_x + 32, top - 3),
                severity,
                fill="#34423B",
                font=small_font,
            )
        draw.text(
            (margin, height - 55),
            (
                f"斑块 {summary.total_patches} 个 · "
                f"受灾 {summary.affected_area_ha:.2f} 公顷"
            ),
            fill="#4C5B53",
            font=label_font,
        )
        output = BytesIO()
        image.save(output, format="PNG", optimize=True)
        return output.getvalue()

    @staticmethod
    def _style_header(sheet: Any, row: int, columns: int) -> None:
        """统一设置工作表表头样式。

        Args:
            sheet: openpyxl 工作表。
            row: 表头行号。
            columns: 表头列数。

        Returns:
            None: 原位设置样式。
        """
        for cell in sheet[row][:columns]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="426A55")
            cell.alignment = Alignment(horizontal="center")

    @classmethod
    def _build_workbook(
        cls,
        task: object,
        summary: DisasterSummaryResponse,
        title: str,
        generated_at: datetime,
        generator: object,
        comment: str,
    ) -> tuple[bytes, dict]:
        """构建包含分布图、统计图和明细的 XLSX 实体。

        Args:
            task: 当前任务。
            summary: 已复核灾害汇总。
            title: 报告标题。
            generated_at: 生成时间。
            generator: 稳定项目用户。
            comment: 生成依据。

        Returns:
            tuple[bytes, dict]: XLSX 字节和可审计生成清单。
        """
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "报告摘要"
        summary_rows = [
            ("报告标题", title),
            ("任务编号", task.task_code),
            ("任务名称", task.task_name),
            ("行政范围", task.administrative_region),
            ("生成时间", generated_at.isoformat()),
            ("生成人", generator.display_name),
            ("用户编码", generator.user_code),
            ("角色快照", generator.role_code),
            ("纳入斑块数", summary.total_patches),
            ("确认斑块数", summary.confirmed_count),
            ("排除斑块数", sum(item.status == "excluded" for item in summary.items)),
            ("受灾面积（公顷）", summary.affected_area_ha),
            ("生成依据", comment),
        ]
        for row in summary_rows:
            summary_sheet.append(row)
        summary_sheet.column_dimensions["A"].width = 24
        summary_sheet.column_dimensions["B"].width = 72
        map_bytes = cls._render_distribution_map(summary)
        map_buffer = BytesIO(map_bytes)
        map_image = SpreadsheetImage(map_buffer)
        map_image.width = 980
        map_image.height = 630
        summary_sheet.add_image(map_image, "A16")

        severity_sheet = workbook.create_sheet("等级统计")
        severity_sheet.append(("灾害等级", "斑块数", "面积（公顷）", "面积占比（%）"))
        for item in summary.by_severity:
            severity_sheet.append(
                (item.label, item.patch_count, item.area_ha, item.percentage)
            )
        cls._style_header(severity_sheet, 1, 4)
        if summary.by_severity:
            pie = PieChart()
            pie.title = "灾害等级面积占比"
            pie.add_data(
                Reference(
                    severity_sheet,
                    min_col=3,
                    min_row=1,
                    max_row=severity_sheet.max_row,
                ),
                titles_from_data=True,
            )
            pie.set_categories(
                Reference(
                    severity_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=severity_sheet.max_row,
                )
            )
            severity_sheet.add_chart(pie, "F2")

        type_sheet = workbook.create_sheet("类型统计")
        type_sheet.append(("灾害类型", "斑块数", "面积（公顷）", "面积占比（%）"))
        for item in summary.by_type:
            type_sheet.append(
                (item.label, item.patch_count, item.area_ha, item.percentage)
            )
        cls._style_header(type_sheet, 1, 4)
        if summary.by_type:
            bar = BarChart()
            bar.title = "灾害类型受灾面积"
            bar.y_axis.title = "公顷"
            bar.add_data(
                Reference(type_sheet, min_col=3, min_row=1, max_row=type_sheet.max_row),
                titles_from_data=True,
            )
            bar.set_categories(
                Reference(type_sheet, min_col=1, min_row=2, max_row=type_sheet.max_row)
            )
            type_sheet.add_chart(bar, "F2")

        detail_sheet = workbook.create_sheet("斑块明细")
        detail_sheet.append(
            (
                "斑块编号",
                "灾害类型",
                "等级",
                "面积（公顷）",
                "作物",
                "识别日期",
                "状态",
                "模型来源",
                "来源版本",
                "来源要素",
                "来源SHA256",
                "复核人",
                "复核依据",
            )
        )
        for item in summary.items:
            detail_sheet.append(
                (
                    item.patch_code,
                    item.disaster_type,
                    item.severity,
                    item.affected_area_ha,
                    item.crop_type or "",
                    item.detected_at.isoformat(),
                    item.status,
                    item.source,
                    item.source_version or "",
                    item.source_feature_id or "",
                    item.source_checksum_sha256 or "",
                    item.reviewed_by or "",
                    item.review_comment or "",
                )
            )
        cls._style_header(detail_sheet, 1, 13)
        detail_sheet.freeze_panes = "A2"
        detail_sheet.auto_filter.ref = detail_sheet.dimensions
        for column in detail_sheet.columns:
            detail_sheet.column_dimensions[column[0].column_letter].width = 18

        source_checksums = sorted(
            {
                item.source_checksum_sha256
                for item in summary.items
                if item.source_checksum_sha256
            }
        )
        manifest = {
            "renderer": "agriscope-disaster-report-xlsx-v1",
            "task_code": task.task_code,
            "generated_at": generated_at.isoformat(),
            "source_patch_count": len(summary.items),
            "confirmed_patch_count": summary.confirmed_count,
            "excluded_patch_count": sum(
                item.status == "excluded" for item in summary.items
            ),
            "affected_area_ha": summary.affected_area_ha,
            "source_checksums": source_checksums,
            "severity_colors": SEVERITY_COLORS,
            "map_png_size_bytes": len(map_bytes),
            "worksheets": [sheet.title for sheet in workbook.worksheets],
        }
        output = BytesIO()
        workbook.save(output)
        return output.getvalue(), manifest

    def verify_report_file(self, report: DisasterReport) -> Path:
        """下载或归档前校验报告路径、签名、大小和 SHA-256。

        Args:
            report: 灾害专题报告模型。

        Returns:
            Path: 已验证 XLSX 实体路径。
        """
        prefix = "storage://disaster-reports/"
        if not report.file_uri.startswith(prefix):
            raise ValidationException("灾害专题报告未登记受控文件地址")
        path = self._resolve_path(report.file_uri.removeprefix(prefix))
        if not path.is_file():
            raise NotFoundException("灾害专题报告实体文件不存在")
        if path.stat().st_size != report.file_size_bytes:
            raise ValidationException("灾害专题报告实体大小校验失败")
        if path.read_bytes()[:4] != b"PK\x03\x04":
            raise ValidationException("灾害专题报告不是有效 XLSX 文件")
        if calculate_sha256(path) != report.checksum_sha256:
            raise ValidationException("灾害专题报告 SHA-256 校验失败")
        return path

    def _to_response(self, report: DisasterReport) -> DisasterReportResponse:
        """转换报告响应并计算当前实体有效状态。

        Args:
            report: 报告模型。

        Returns:
            DisasterReportResponse: 前端报告摘要。
        """
        effective_status = report.status
        try:
            self.verify_report_file(report)
        except (NotFoundException, ValidationException):
            effective_status = "invalid"
        is_current = effective_status == "completed"
        return DisasterReportResponse(
            report_code=report.report_code,
            report_title=report.report_title,
            status=effective_status,
            file_size_bytes=report.file_size_bytes,
            checksum_sha256=report.checksum_sha256,
            source_patch_count=report.source_patch_count,
            source_confirmed_count=report.source_confirmed_count,
            source_excluded_count=report.source_excluded_count,
            source_latest_updated_at=report.source_latest_updated_at,
            affected_area_ha=float(report.affected_area_ha),
            report_manifest=report.report_manifest,
            generation_comment=report.generation_comment,
            generated_by=report.generated_by,
            generated_by_code=report.generated_by_code,
            generated_by_role=report.generated_by_role,
            generated_at=report.generated_at,
            download_url=(
                f"/api/v1/disasters/reports/{report.report_code}/download"
                if effective_status != "invalid"
                else None
            ),
            is_current=is_current,
        )

    async def list_reports(
        self,
        db: AsyncSession,
        task_code: str,
    ) -> DisasterReportListResponse:
        """查询任务灾害专题报告及实体状态。

        Args:
            db: 异步数据库会话。
            task_code: 作业任务编号。

        Returns:
            DisasterReportListResponse: 报告列表。
        """
        task = await self.workbench_dao.get_task_by_code(db, task_code)
        if task is None:
            raise NotFoundException(f"未找到任务 {task_code}")
        reports = await self.dao.list_reports(db, task.id)
        items = await asyncio.gather(
            *[asyncio.to_thread(self._to_response, report) for report in reports]
        )
        return DisasterReportListResponse(task_code=task_code, items=list(items))

    async def generate_report(
        self,
        db: AsyncSession,
        task_code: str,
        request: DisasterReportGenerateRequest,
    ) -> DisasterReportResponse:
        """通过复核门禁后生成不可变 XLSX 灾害专题报告。

        Args:
            db: 异步数据库会话。
            task_code: 作业任务编号。
            request: 操作人、标题和生成依据。

        Returns:
            DisasterReportResponse: 新生成报告。
        """
        task = await self.workbench_dao.get_task_by_code_for_update(db, task_code)
        if task is None:
            raise NotFoundException(f"未找到任务 {task_code}")
        if task.status == "completed":
            raise ValidationException("已完成任务不得重新生成灾害专题报告")
        operator = await self.project_user_service.require_capability(
            db,
            task.project_id,
            request.operator_code,
            "generate_disaster_report",
        )
        rows = await self.disaster_dao.get_patches(db, task.id)
        if not rows:
            raise ValidationException("尚未导入灾害模型结果，不能生成专题报告")
        pending_count = sum(patch.status == "pending" for patch, _ in rows)
        if pending_count:
            raise ValidationException(
                f"仍有 {pending_count} 个灾害斑块待复核，不能生成专题报告"
            )
        summary = await self.disaster_service.get_summary(db, task_code)
        generated_at = datetime.now(UTC)
        source_latest_updated_at = max(patch.updated_at for patch, _ in rows)
        report_code = (
            f"DSRPT-{task_code}-{generated_at:%Y%m%dT%H%M%S}-"
            f"{secrets.token_hex(4)}"
        )
        content, manifest = await asyncio.to_thread(
            self._build_workbook,
            task,
            summary,
            request.report_title,
            generated_at,
            operator,
            request.comment,
        )
        relative_path = f"{task_code}/{report_code}.xlsx"
        final_path = self._resolve_path(relative_path)
        temporary_path = self._resolve_path(
            f"{task_code}/.{report_code}.{secrets.token_hex(6)}.tmp"
        )
        final_path.parent.mkdir(parents=True, exist_ok=True)
        await asyncio.to_thread(temporary_path.write_bytes, content)
        await asyncio.to_thread(os.replace, temporary_path, final_path)
        report = DisasterReport(
            task_id=task.id,
            report_code=report_code,
            report_title=request.report_title,
            status="completed",
            file_uri=f"storage://disaster-reports/{relative_path}",
            file_size_bytes=final_path.stat().st_size,
            checksum_sha256=calculate_sha256(final_path),
            source_patch_count=len(rows),
            source_confirmed_count=summary.confirmed_count,
            source_excluded_count=sum(
                patch.status == "excluded" for patch, _ in rows
            ),
            source_latest_updated_at=source_latest_updated_at,
            affected_area_ha=summary.affected_area_ha,
            report_manifest=manifest,
            generation_comment=request.comment,
            generated_by=operator.display_name,
            generated_by_code=operator.user_code,
            generated_by_role=operator.role_code,
            generated_at=generated_at,
        )
        try:
            superseded_count = await self.dao.supersede_completed_reports(
                db,
                task.id,
            )
            report = await self.dao.add_report(db, report)
            task.updated_at = generated_at
            await self.workbench_dao.add_review_record(
                db,
                ReviewRecord(
                    task_id=task.id,
                    review_level="disaster_monitoring",
                    action="disaster_report_generated",
                    reviewer=operator.display_name,
                    reviewer_code=operator.user_code,
                    reviewer_role=operator.role_code,
                    comment=(
                        f"生成灾害专题报告 {report_code}；"
                        f"斑块 {len(rows)} 个，确认 {summary.confirmed_count} 个，"
                        f"排除 {report.source_excluded_count} 个；"
                        f"替代历史报告 {superseded_count} 份；{request.comment}"
                    ),
                ),
            )
            await db.commit()
        except SQLAlchemyError:
            await db.rollback()
            await asyncio.to_thread(final_path.unlink, missing_ok=True)
            raise
        return self._to_response(report)

    async def authorize_download(
        self,
        db: AsyncSession,
        report_code: str,
        requester_code: str,
    ) -> DisasterReportDownload:
        """鉴权、复核并审计灾害专题报告下载。

        Args:
            db: 异步数据库会话。
            report_code: 报告编号。
            requester_code: 下载人稳定用户编码。

        Returns:
            DisasterReportDownload: 已验证下载信息。
        """
        report = await self.dao.get_report_by_code(db, report_code)
        if report is None:
            raise NotFoundException(f"未找到灾害专题报告 {report_code}")
        task = await self.workbench_dao.get_task_by_id(db, report.task_id)
        if task is None:
            raise NotFoundException("灾害专题报告关联任务不存在")
        requester = await self.project_user_service.require_capability(
            db,
            task.project_id,
            requester_code,
            "download_disaster_report",
        )
        path = await asyncio.to_thread(self.verify_report_file, report)
        await self.workbench_dao.add_review_record(
            db,
            ReviewRecord(
                task_id=task.id,
                review_level="disaster_monitoring",
                action="disaster_report_downloaded",
                reviewer=requester.display_name,
                reviewer_code=requester.user_code,
                reviewer_role=requester.role_code,
                comment=(
                    f"下载灾害专题报告 {report.report_code}；"
                    f"SHA256 {report.checksum_sha256}"
                ),
            ),
        )
        await db.commit()
        return DisasterReportDownload(
            path=path,
            filename=f"{report.report_code}.xlsx",
            checksum_sha256=report.checksum_sha256,
        )
