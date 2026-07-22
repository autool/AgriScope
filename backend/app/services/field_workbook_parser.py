"""外业核查 Excel 工作簿安全解析与模板生成。"""

from io import BytesIO
from pathlib import PurePosixPath
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from app.core.exceptions import ValidationException
from app.schemas.field_verification import FieldVerificationImportItem

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
WORKSHEET_NAME = "外业核查导入"
MAX_XLSX_BYTES = 10 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 2000
MAX_IMPORT_RECORDS = 500
REQUIRED_HEADERS = (
    "verification_code",
    "source_record_id",
    "lon",
    "lat",
    "observed_land_class",
    "observed_crop_type",
    "captured_at",
    "photo_urls",
)
OPTIONAL_HEADERS = ("voice_url", "remark")
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS


class FieldVerificationWorkbookParser:
    """解析受控 `.xlsx` 外业核查文件并生成标准模板。"""

    @staticmethod
    def _validate_archive(content: bytes) -> None:
        """校验 XLSX ZIP 结构、路径和解压后体积。

        Args:
            content: 原始 XLSX 字节。

        Returns:
            None: 校验通过时无返回值。
        """
        try:
            with ZipFile(BytesIO(content)) as archive:
                entries = archive.infolist()
                if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
                    raise ValidationException("Excel 文件内部条目数量异常")
                if "[Content_Types].xml" not in archive.namelist():
                    raise ValidationException("Excel 文件缺少标准 XLSX 内容类型声明")
                uncompressed_size = 0
                for entry in entries:
                    path = PurePosixPath(entry.filename)
                    if (
                        path.is_absolute()
                        or ".." in path.parts
                        or "\\" in entry.filename
                    ):
                        raise ValidationException("Excel 文件包含不安全的内部路径")
                    if entry.flag_bits & 0x1:
                        raise ValidationException("不支持加密的 Excel 文件")
                    uncompressed_size += entry.file_size
                    if uncompressed_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                        raise ValidationException("Excel 文件解压后不得超过 50MB")
        except BadZipFile as exc:
            raise ValidationException("Excel 文件不是有效的 XLSX 压缩包") from exc

    @staticmethod
    def _cell_text(value: object) -> str:
        """将 Excel 单元格转换为稳定文本。

        Args:
            value: openpyxl 单元格值。

        Returns:
            str: 去除首尾空白的文本。
        """
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if hasattr(value, "isoformat"):
            return str(value.isoformat())
        return str(value).strip()

    @staticmethod
    def _validation_message(error: ValidationError) -> str:
        """提取 Pydantic 首条安全校验信息。

        Args:
            error: 单条外业记录校验异常。

        Returns:
            str: 可展示的中文错误信息。
        """
        first_error = error.errors(include_url=False)[0]
        location = ".".join(str(item) for item in first_error.get("loc", ()))
        message = str(first_error.get("msg") or "记录格式不合法")
        message = message.removeprefix("Value error, ")
        return f"{location}: {message}" if location else message

    def parse(self, filename: str, content: bytes) -> list[FieldVerificationImportItem]:
        """解析 Excel 首个业务工作表为外业核查记录。

        Args:
            filename: 上传文件名。
            content: 原始 XLSX 字节。

        Returns:
            list[FieldVerificationImportItem]: 已完成业务校验的记录。
        """
        normalized_filename = filename.strip()
        if not normalized_filename.lower().endswith(".xlsx"):
            raise ValidationException("仅支持 .xlsx 外业核查文件")
        if not content:
            raise ValidationException("Excel 文件内容为空")
        if len(content) > MAX_XLSX_BYTES:
            raise ValidationException("Excel 文件不得超过 10MB")
        self._validate_archive(content)
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ValidationException("无法读取 Excel 工作簿") from exc
        try:
            worksheet = (
                workbook[WORKSHEET_NAME]
                if WORKSHEET_NAME in workbook.sheetnames
                else workbook.active
            )
            rows = worksheet.iter_rows()
            header_cells = next(rows, None)
            if header_cells is None:
                raise ValidationException("Excel 必须包含表头和至少一条记录")
            if any(cell.data_type == "f" for cell in header_cells):
                raise ValidationException("Excel 表头不得包含公式")
            headers = [self._cell_text(cell.value) for cell in header_cells]
            if len(headers) != len(set(headers)):
                raise ValidationException("Excel 表头不得重复")
            missing_header = next(
                (header for header in REQUIRED_HEADERS if header not in headers),
                None,
            )
            if missing_header:
                raise ValidationException(f"Excel 缺少必填列 {missing_header}")
            header_index = {header: index for index, header in enumerate(headers)}
            records: list[FieldVerificationImportItem] = []
            for row_number, row in enumerate(rows, start=2):
                if any(cell.data_type == "f" for cell in row):
                    raise ValidationException(
                        f"Excel 第 {row_number} 行包含公式，必须粘贴为值后再导入"
                    )
                values = [self._cell_text(cell.value) for cell in row]
                if not any(values):
                    continue
                if len(records) >= MAX_IMPORT_RECORDS:
                    raise ValidationException("单次最多导入 500 条外业记录")

                row_values = {
                    header: values[index] if index < len(values) else ""
                    for header, index in header_index.items()
                }

                photo_urls = [
                    item.strip()
                    for item in row_values.get("photo_urls", "").split("|")
                    if item.strip()
                ]
                try:
                    records.append(
                        FieldVerificationImportItem.model_validate(
                            {
                                "verification_code": row_values.get(
                                    "verification_code", ""
                                ),
                                "source_record_id": row_values.get(
                                    "source_record_id", ""
                                ),
                                "lon": row_values.get("lon", ""),
                                "lat": row_values.get("lat", ""),
                                "observed_land_class": (
                                    row_values.get("observed_land_class") or None
                                ),
                                "observed_crop_type": (
                                    row_values.get("observed_crop_type") or None
                                ),
                                "captured_at": row_values.get("captured_at", ""),
                                "photo_urls": photo_urls,
                                "voice_url": row_values.get("voice_url") or None,
                                "remark": row_values.get("remark") or None,
                            }
                        )
                    )
                except ValidationError as exc:
                    raise ValidationException(
                        f"Excel 第 {row_number} 行 {self._validation_message(exc)}"
                    ) from exc
            if not records:
                raise ValidationException("Excel 必须包含至少一条外业核查记录")
            return records
        finally:
            workbook.close()

    @staticmethod
    def build_template() -> bytes:
        """生成带字段说明和地类下拉选项的 XLSX 模板。

        Returns:
            bytes: 可直接下载的标准 Excel 工作簿。
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = WORKSHEET_NAME
        worksheet.append(list(ALL_HEADERS))
        worksheet.append(
            [
                "FV-2026-0001",
                "mobile-record-0001",
                126.45,
                45.75,
                "耕地",
                "玉米",
                "2026-07-22T09:30:00+08:00",
                (
                    "https://files.example/field/0001-1.jpg|"
                    "https://files.example/field/0001-2.jpg"
                ),
                "https://files.example/field/0001.m4a",
                "现场边界清晰",
            ]
        )
        header_fill = PatternFill("solid", fgColor="245A3F")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:J{MAX_IMPORT_RECORDS + 1}"
        widths = [20, 24, 14, 14, 20, 20, 30, 52, 42, 36]
        for column, width in zip("ABCDEFGHIJ", widths, strict=True):
            worksheet.column_dimensions[column].width = width
        for cell in worksheet[2]:
            cell.number_format = "@"
        land_class_validation = DataValidation(
            type="list",
            formula1='"耕地,园地,林地,草地,水域,建设用地"',
            allow_blank=True,
        )
        worksheet.add_data_validation(land_class_validation)
        land_class_validation.add(f"E2:E{MAX_IMPORT_RECORDS + 1}")

        instructions = workbook.create_sheet("填写说明")
        instructions.append(["字段", "要求"])
        instructions.append(
            [
                "verification_code",
                "任务内唯一，只允许字母、数字、下划线和短横线",
            ]
        )
        instructions.append(["source_record_id", "外业采集系统中的唯一记录编号"])
        instructions.append(["lon / lat", "WGS84（EPSG:4326）经纬度"])
        instructions.append(
            [
                "observed_land_class",
                "耕地、园地、林地、草地、水域或建设用地",
            ]
        )
        instructions.append(["observed_crop_type", "耕地必填，非耕地必须为空"])
        instructions.append(["captured_at", "必须填写带时区的 ISO 8601 文本"])
        instructions.append(["photo_urls", "至少一个受控照片 URI，多个使用 | 分隔"])
        instructions.append(["voice_url / remark", "可选语音 URI和现场备注"])
        instructions.column_dimensions["A"].width = 26
        instructions.column_dimensions["B"].width = 72
        for cell in instructions[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()
