"""任务地块属性 Excel 安全解析、字段模式固化与导出生成。"""

import json
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
from pathlib import PurePosixPath
from typing import overload
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from app.core.exceptions import ValidationException
from app.models.plot import FarmlandPlot
from app.schemas.plot_attribute_field import PlotAttributeFieldDefinition
from app.schemas.plot_attribute_workbook import PlotAttributeWorkbookRow

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
WORKSHEET_NAME = "地块属性"
SCHEMA_SHEET_NAME = "_schema"
SCHEMA_VERSION = "plot-attributes-v2"
MAX_XLSX_BYTES = 10 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 2000
MAX_WORKBOOK_ROWS = 500
BASE_WORKBOOK_HEADERS = (
    "plot_code",
    "expected_version",
    "owner_village",
    "land_class",
    "crop_type",
    "planting_mode",
    "irrigation_condition",
)
# 保留固定字段常量名，供既有校验与外部脚本读取；动态列由字段模式追加。
WORKBOOK_HEADERS = BASE_WORKBOOK_HEADERS
EMPTY_SCHEMA_DIGEST = sha256(b"[]").hexdigest()


@dataclass(frozen=True)
class ParsedPlotAttributeWorkbook(Sequence[PlotAttributeWorkbookRow]):
    """已校验工作簿行及其历史字段模式。"""

    rows: list[PlotAttributeWorkbookRow]
    definition_snapshot: list[dict]
    definition_digest: str

    def __len__(self) -> int:
        """返回工作簿记录数量。

        Returns:
            int: 记录数量。
        """
        return len(self.rows)

    def __iter__(self) -> Iterator[PlotAttributeWorkbookRow]:
        """迭代工作簿记录。

        Returns:
            Iterator[PlotAttributeWorkbookRow]: 记录迭代器。
        """
        return iter(self.rows)

    @overload
    def __getitem__(self, index: int) -> PlotAttributeWorkbookRow: ...

    @overload
    def __getitem__(self, index: slice) -> list[PlotAttributeWorkbookRow]: ...

    def __getitem__(
        self,
        index: int | slice,
    ) -> PlotAttributeWorkbookRow | list[PlotAttributeWorkbookRow]:
        """按索引读取记录。

        Args:
            index: 整数索引或切片。

        Returns:
            PlotAttributeWorkbookRow | list[PlotAttributeWorkbookRow]: 记录。
        """
        return self.rows[index]


class PlotAttributeWorkbookParser:
    """解析任务地块属性 XLSX，并生成模式锁定的标准工作簿。"""

    @staticmethod
    def schema_digest(snapshot: list[dict]) -> str:
        """计算字段模式规范化 SHA-256。

        Args:
            snapshot: 字段模式快照。

        Returns:
            str: 64 位十六进制摘要。
        """
        canonical = json.dumps(
            snapshot,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        return sha256(canonical).hexdigest()

    @staticmethod
    def _validate_archive(content: bytes) -> None:
        """校验 XLSX ZIP 路径、加密标记和解压规模。

        Args:
            content: 原始 XLSX 字节。

        Returns:
            None: 结构安全时无返回值。
        """
        try:
            with ZipFile(BytesIO(content)) as archive:
                entries = archive.infolist()
                if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
                    raise ValidationException("Excel 文件内部条目数量异常")
                required = {"[Content_Types].xml", "xl/workbook.xml"}
                if not required.issubset(archive.namelist()):
                    raise ValidationException("Excel 文件缺少标准 XLSX 结构")
                expanded_size = 0
                for entry in entries:
                    path = PurePosixPath(entry.filename)
                    if (
                        path.is_absolute()
                        or ".." in path.parts
                        or "\\" in entry.filename
                        or not entry.filename
                    ):
                        raise ValidationException("Excel 文件包含不安全的内部路径")
                    if entry.flag_bits & 0x1:
                        raise ValidationException("不支持加密的 Excel 文件")
                    expanded_size += entry.file_size
                    if expanded_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                        raise ValidationException("Excel 文件解压后不得超过 50MB")
                if archive.testzip() is not None:
                    raise ValidationException("Excel 压缩内容已损坏")
        except BadZipFile as exc:
            raise ValidationException("Excel 文件不是有效的 XLSX 压缩包") from exc

    @staticmethod
    def _cell_text(value: object) -> str:
        """将单元格值转换为稳定文本。

        Args:
            value: openpyxl 单元格值。

        Returns:
            str: 去除首尾空白的文本。
        """
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _validation_message(error: ValidationError) -> str:
        """提取首条可安全展示的 Pydantic 错误。

        Args:
            error: 单行属性校验异常。

        Returns:
            str: 字段定位和中文错误信息。
        """
        first_error = error.errors(include_url=False)[0]
        location = ".".join(str(item) for item in first_error.get("loc", ()))
        message = str(first_error.get("msg") or "记录格式不合法")
        message = message.removeprefix("Value error, ")
        return f"{location}: {message}" if location else message

    @classmethod
    def _read_schema(
        cls,
        workbook: object,
    ) -> tuple[list[PlotAttributeFieldDefinition], list[dict], str]:
        """读取隐藏字段模式并校验其自身摘要。

        Args:
            workbook: 已打开的 openpyxl 工作簿。

        Returns:
            tuple[list[PlotAttributeFieldDefinition], list[dict], str]:
                类型化定义、JSON 快照和摘要。
        """
        if SCHEMA_SHEET_NAME not in workbook.sheetnames:
            return [], [], EMPTY_SCHEMA_DIGEST
        schema_sheet = workbook[SCHEMA_SHEET_NAME]
        payload = {
            cls._cell_text(row[0].value): row[1].value
            for row in schema_sheet.iter_rows(min_row=1, max_col=2)
            if row and cls._cell_text(row[0].value)
        }
        if cls._cell_text(payload.get("schema_version")) != SCHEMA_VERSION:
            raise ValidationException("Excel 自定义字段模式版本不受支持")
        digest = cls._cell_text(payload.get("definition_digest"))
        raw_snapshot = cls._cell_text(payload.get("definition_snapshot_json"))
        try:
            parsed_snapshot = json.loads(raw_snapshot)
            if not isinstance(parsed_snapshot, list):
                raise TypeError
            definitions = [
                PlotAttributeFieldDefinition.model_validate(item)
                for item in parsed_snapshot
            ]
        except (TypeError, ValueError, ValidationError, json.JSONDecodeError) as exc:
            raise ValidationException("Excel 自定义字段模式快照不合法") from exc
        snapshot = [item.model_dump(mode="json") for item in definitions]
        if digest != cls.schema_digest(snapshot):
            raise ValidationException("Excel 自定义字段模式摘要校验失败")
        codes = [item.field_code for item in definitions]
        if len(codes) != len(set(codes)):
            raise ValidationException("Excel 自定义字段模式包含重复编码")
        return definitions, snapshot, digest

    @staticmethod
    def _custom_cell_value(
        definition: PlotAttributeFieldDefinition,
        value: object,
        row_number: int,
    ) -> object:
        """按历史字段定义解析一个自定义属性单元格。

        Args:
            definition: 工作簿内固化字段定义。
            value: 原始单元格值。
            row_number: Excel 行号。

        Returns:
            object: 标准 JSON 标量。
        """
        label = definition.label
        if value is None or (isinstance(value, str) and not value.strip()):
            if definition.required:
                raise ValidationException(
                    f"Excel 第 {row_number} 行自定义字段“{label}”为必填项"
                )
            return None
        if definition.field_type == "text":
            normalized = str(value).strip()
            if len(normalized) > 500:
                raise ValidationException(
                    f"Excel 第 {row_number} 行自定义字段“{label}”过长"
                )
            return normalized
        if definition.field_type == "number":
            if isinstance(value, bool):
                raise ValidationException(
                    f"Excel 第 {row_number} 行自定义字段“{label}”必须为数值"
                )
            try:
                numeric = float(value)
            except (TypeError, ValueError) as exc:
                raise ValidationException(
                    f"Excel 第 {row_number} 行自定义字段“{label}”必须为数值"
                ) from exc
            if numeric != numeric or numeric in {float("inf"), float("-inf")}:
                raise ValidationException(
                    f"Excel 第 {row_number} 行自定义字段“{label}”必须为有限数值"
                )
            return int(numeric) if numeric.is_integer() else numeric
        if definition.field_type == "date":
            if isinstance(value, datetime):
                return value.date().isoformat()
            if isinstance(value, date):
                return value.isoformat()
            try:
                return date.fromisoformat(str(value).strip()).isoformat()
            except ValueError as exc:
                raise ValidationException(
                    f"Excel 第 {row_number} 行自定义字段“{label}”"
                    "必须使用 YYYY-MM-DD 日期格式"
                ) from exc
        if definition.field_type == "boolean":
            if isinstance(value, bool):
                return value
            normalized = str(value).strip().lower()
            if normalized in {"是", "true", "1"}:
                return True
            if normalized in {"否", "false", "0"}:
                return False
            raise ValidationException(
                f"Excel 第 {row_number} 行自定义字段“{label}”必须填写是或否"
            )
        normalized = str(value).strip()
        if normalized not in definition.options:
            raise ValidationException(
                f"Excel 第 {row_number} 行自定义字段“{label}”不在选项范围内"
            )
        return normalized

    def parse(
        self,
        filename: str,
        content: bytes,
        *,
        expected_snapshot: list[dict] | None = None,
        expected_digest: str | None = None,
    ) -> ParsedPlotAttributeWorkbook:
        """解析严格表头和历史字段模式锁定的属性工作簿。

        Args:
            filename: 原始上传文件名。
            content: 原始 XLSX 字节。
            expected_snapshot: 可选服务端历史模式快照。
            expected_digest: 可选服务端历史模式摘要。

        Returns:
            ParsedPlotAttributeWorkbook: 完成结构、模式和逐行业务校验的结果。
        """
        normalized_filename = filename.strip()
        if not normalized_filename.lower().endswith(".xlsx"):
            raise ValidationException("仅支持 .xlsx 地块属性工作簿")
        if not content:
            raise ValidationException("Excel 文件内容为空")
        if len(content) > MAX_XLSX_BYTES:
            raise ValidationException("Excel 文件不得超过 10MB")
        self._validate_archive(content)
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ValidationException("无法读取 Excel 工作簿") from exc
        try:
            if WORKSHEET_NAME not in workbook.sheetnames:
                raise ValidationException(f"Excel 必须包含“{WORKSHEET_NAME}”工作表")
            for sheet in workbook.worksheets:
                for sheet_row in sheet.iter_rows():
                    if any(cell.data_type == "f" for cell in sheet_row):
                        raise ValidationException(
                            f"Excel 工作表“{sheet.title}”包含公式，"
                            "必须粘贴为值后再导入"
                        )
            definitions, snapshot, digest = self._read_schema(workbook)
            if expected_snapshot is not None and snapshot != expected_snapshot:
                raise ValidationException("Excel 字段定义已变化，请重新导出工作簿")
            if expected_digest is not None and digest != expected_digest:
                raise ValidationException("Excel 字段模式已变化，请重新导出工作簿")

            headers_expected = BASE_WORKBOOK_HEADERS + tuple(
                f"custom:{item.field_code}" for item in definitions
            )
            worksheet = workbook[WORKSHEET_NAME]
            rows = worksheet.iter_rows()
            header_cells = next(rows, None)
            if header_cells is None:
                raise ValidationException("Excel 必须包含标准表头和至少一条记录")
            headers = tuple(self._cell_text(cell.value) for cell in header_cells)
            if headers != headers_expected:
                raise ValidationException(
                    "Excel 表头必须严格为：" + "、".join(headers_expected)
                )

            records: list[PlotAttributeWorkbookRow] = []
            seen_codes: set[str] = set()
            for row_number, row in enumerate(rows, start=2):
                raw_values = [cell.value for cell in row]
                if not any(
                    value is not None and str(value).strip() for value in raw_values
                ):
                    continue
                if len(records) >= MAX_WORKBOOK_ROWS:
                    raise ValidationException("单次最多导入 500 条地块属性")
                raw_values.extend([None] * (len(headers_expected) - len(raw_values)))
                fixed_values = {
                    header: self._cell_text(raw_values[index])
                    for index, header in enumerate(BASE_WORKBOOK_HEADERS)
                }
                custom_attributes = {
                    definition.field_code: self._custom_cell_value(
                        definition,
                        raw_values[len(BASE_WORKBOOK_HEADERS) + index],
                        row_number,
                    )
                    for index, definition in enumerate(definitions)
                }
                try:
                    record = PlotAttributeWorkbookRow.model_validate(
                        {
                            "plot_code": fixed_values["plot_code"],
                            "expected_version": fixed_values["expected_version"],
                            "owner_village": fixed_values["owner_village"],
                            "land_class": fixed_values["land_class"],
                            "crop_type": fixed_values["crop_type"] or None,
                            "planting_mode": fixed_values["planting_mode"] or None,
                            "irrigation_condition": (
                                fixed_values["irrigation_condition"] or None
                            ),
                            "custom_attributes": custom_attributes,
                        }
                    )
                except ValidationError as exc:
                    raise ValidationException(
                        f"Excel 第 {row_number} 行 {self._validation_message(exc)}"
                    ) from exc
                if record.plot_code in seen_codes:
                    raise ValidationException(
                        f"Excel 第 {row_number} 行图斑编号 {record.plot_code} 重复"
                    )
                seen_codes.add(record.plot_code)
                records.append(record)
            if not records:
                raise ValidationException("Excel 必须包含至少一条地块属性记录")
            return ParsedPlotAttributeWorkbook(
                rows=records,
                definition_snapshot=snapshot,
                definition_digest=digest,
            )
        finally:
            workbook.close()

    @classmethod
    def build_export(
        cls,
        plots: Sequence[FarmlandPlot],
        definition_snapshot: list[dict] | None = None,
    ) -> bytes:
        """从数据库当前值生成带字段模式快照的可回写 XLSX。

        Args:
            plots: 已按任务范围校验的有效图斑。
            definition_snapshot: 当前活动自定义字段模式快照。

        Returns:
            bytes: 标准地块属性 Excel 实体。
        """
        snapshot = definition_snapshot or []
        definitions = [
            PlotAttributeFieldDefinition.model_validate(item) for item in snapshot
        ]
        digest = cls.schema_digest(snapshot)
        headers = BASE_WORKBOOK_HEADERS + tuple(
            f"custom:{item.field_code}" for item in definitions
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = WORKSHEET_NAME
        worksheet.append(list(headers))
        for plot in plots:
            custom_values = getattr(plot, "custom_attributes", {}) or {}
            values: list[object] = [
                plot.plot_code,
                plot.version,
                plot.owner_village or "",
                plot.land_class or "",
                plot.crop_type or "",
                plot.planting_mode or "",
                plot.irrigation_condition or "",
            ]
            for definition in definitions:
                value = custom_values.get(definition.field_code)
                if definition.field_type == "boolean" and value is not None:
                    value = "是" if value else "否"
                values.append("" if value is None else value)
            worksheet.append(values)

        header_fill = PatternFill("solid", fgColor="245A3F")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        worksheet.freeze_panes = "A2"
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{len(plots) + 1}"
        widths = [24, 18, 24, 18, 18, 20, 20]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width
        for index, definition in enumerate(
            definitions,
            start=len(BASE_WORKBOOK_HEADERS) + 1,
        ):
            worksheet.column_dimensions[get_column_letter(index)].width = max(
                18,
                min(len(definition.label) * 2 + 8, 32),
            )

        land_class_validation = DataValidation(
            type="list",
            formula1='"耕地,园地,林地,草地,水域,建设用地"',
            allow_blank=False,
        )
        worksheet.add_data_validation(land_class_validation)
        land_class_validation.add(f"D2:D{MAX_WORKBOOK_ROWS + 1}")
        single_select_ranges: list[tuple[str, PlotAttributeFieldDefinition]] = []
        for index, definition in enumerate(
            definitions,
            start=len(BASE_WORKBOOK_HEADERS) + 1,
        ):
            column = get_column_letter(index)
            if definition.field_type == "single_select":
                range_name = f"custom_options_{index}"
                validation = DataValidation(
                    type="list",
                    formula1=range_name,
                    allow_blank=not definition.required,
                )
                worksheet.add_data_validation(validation)
                validation.add(f"{column}2:{column}{MAX_WORKBOOK_ROWS + 1}")
                single_select_ranges.append((range_name, definition))
            elif definition.field_type == "boolean":
                validation = DataValidation(
                    type="list",
                    formula1='"是,否"',
                    allow_blank=not definition.required,
                )
                worksheet.add_data_validation(validation)
                validation.add(f"{column}2:{column}{MAX_WORKBOOK_ROWS + 1}")
            elif definition.field_type == "date":
                for cell in worksheet[column][1:]:
                    cell.number_format = "yyyy-mm-dd"

        instructions = workbook.create_sheet("填写说明")
        instructions.append(["字段", "填写要求"])
        instructions.append(["plot_code", "只读业务主键，不得修改或重复"])
        instructions.append(
            ["expected_version", "只读并发版本；导入时必须与数据库当前版本一致"]
        )
        instructions.append(["owner_village", "必填，最长 100 个字符"])
        instructions.append(["land_class", "耕地、园地、林地、草地、水域或建设用地"])
        instructions.append(["crop_type", "耕地必填，非耕地必须为空"])
        for definition in definitions:
            detail = {
                "text": "文本，最长 500 个字符",
                "number": "有限数值",
                "date": "日期 YYYY-MM-DD",
                "boolean": "是或否",
                "single_select": "选项：" + "、".join(definition.options),
            }[definition.field_type]
            required = "必填" if definition.required else "可选"
            instructions.append(
                [
                    f"custom:{definition.field_code}",
                    f"{definition.label}；{required}；{detail}",
                ]
            )
        instructions.append(
            [
                "原子导入",
                "任一图斑越权、重复、版本过期、字段模式变化或属性不合法时整批不写入",
            ]
        )
        instructions.column_dimensions["A"].width = 28
        instructions.column_dimensions["B"].width = 88
        for cell in instructions[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        schema_sheet = workbook.create_sheet(SCHEMA_SHEET_NAME)
        schema_sheet.append(["schema_version", SCHEMA_VERSION])
        schema_sheet.append(["definition_digest", digest])
        schema_sheet.append(
            [
                "definition_snapshot_json",
                json.dumps(
                    snapshot,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                ),
            ]
        )
        for index, (range_name, definition) in enumerate(
            single_select_ranges,
            start=4,
        ):
            column = get_column_letter(index)
            schema_sheet.cell(row=1, column=index, value=range_name)
            for row_index, option in enumerate(definition.options, start=2):
                schema_sheet.cell(row=row_index, column=index, value=option)
            workbook.defined_names.add(
                DefinedName(
                    range_name,
                    attr_text=(
                        f"'{SCHEMA_SHEET_NAME}'!${column}$2:"
                        f"${column}${len(definition.options) + 1}"
                    ),
                )
            )
        schema_sheet.sheet_state = "veryHidden"

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()
