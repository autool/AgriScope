"""面积统计 XLSX、PDF 和清单实体渲染器。"""

import json
import math
from datetime import datetime
from hashlib import sha256
from io import BytesIO
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont

from app.schemas.statistics import AreaGroupItem, AreaStatisticsResponse

REPORT_COLORS = (
    "#347C57",
    "#D89A3C",
    "#4F7FAF",
    "#8C6AAA",
    "#C95F4B",
    "#5B9A96",
    "#7A8D55",
    "#A97045",
)


class StatisticsReportRenderer:
    """将可信面积统计快照渲染为正式报告实体。"""

    @staticmethod
    def _load_font(
        size: int,
    ) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        """加载可显示中文的系统字体。

        Args:
            size: 字体像素大小。

        Returns:
            ImageFont.FreeTypeFont | ImageFont.ImageFont: 可用字体。
        """
        candidates = (
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
        )
        for candidate in candidates:
            try:
                return ImageFont.truetype(candidate, size=size)
            except OSError:
                continue
        return ImageFont.load_default()

    @staticmethod
    def _style_header(sheet: Any, row: int, columns: int) -> None:
        """设置工作表表头样式。

        Args:
            sheet: openpyxl 工作表。
            row: 表头行号。
            columns: 表头列数。

        Returns:
            None: 原位修改样式。
        """
        for cell in sheet[row][:columns]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="426A55")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @classmethod
    def _add_group_sheet(
        cls,
        workbook: Workbook,
        title: str,
        items: list[AreaGroupItem],
        *,
        chart_kind: str | None = None,
    ) -> None:
        """新增一个统计分组工作表并按需添加图表。

        Args:
            workbook: 目标工作簿。
            title: 工作表名称。
            items: 分组统计项。
            chart_kind: `bar`、`pie` 或空。

        Returns:
            None: 工作簿原位增加工作表。
        """
        sheet = workbook.create_sheet(title)
        sheet.append(
            [
                "编码",
                "名称",
                "上级区域",
                "图斑数",
                "面积（公顷）",
                "面积（亩）",
                "占比（%）",
            ]
        )
        cls._style_header(sheet, 1, 7)
        for item in items:
            sheet.append(
                [
                    item.code or "",
                    item.label,
                    item.parent_label or "",
                    item.plot_count,
                    item.area_ha,
                    item.area_mu,
                    item.percentage,
                ]
            )
        widths = (16, 24, 24, 12, 16, 16, 14)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.freeze_panes = "A2"
        if not items or chart_kind is None:
            return
        max_row = min(len(items) + 1, 16)
        labels = Reference(sheet, min_col=2, min_row=2, max_row=max_row)
        values = Reference(sheet, min_col=5, min_row=1, max_row=max_row)
        if chart_kind == "pie":
            chart = PieChart()
            chart.title = f"{title}面积占比"
        else:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = f"{title}面积排名"
            chart.y_axis.title = title
            chart.x_axis.title = "面积（公顷）"
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 8
        chart.width = 15
        sheet.add_chart(chart, "I2")

    @classmethod
    def build_xlsx(
        cls,
        task: object,
        summary: AreaStatisticsResponse,
        report_title: str,
        generated_at: datetime,
        operator: object,
        comment: str,
    ) -> bytes:
        """生成包含六维统计、年度趋势和图表的 XLSX。

        Args:
            task: 当前作业任务。
            summary: 任务实时统计快照。
            report_title: 报告标题。
            generated_at: 生成时间。
            operator: 持久化项目用户。
            comment: 报告生成依据。

        Returns:
            bytes: XLSX 实体字节。
        """
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "报告摘要"
        summary_sheet.append([report_title])
        summary_sheet.merge_cells("A1:D1")
        summary_sheet["A1"].font = Font(size=18, bold=True, color="244E39")
        summary_sheet["A1"].alignment = Alignment(horizontal="center")
        summary_rows = (
            ("任务编号", task.task_code),
            ("任务名称", task.task_name),
            ("监测年度", summary.monitor_year),
            ("统计时间", summary.generated_at.isoformat()),
            ("报告生成时间", generated_at.isoformat()),
            ("生成用户", operator.display_name),
            ("用户编码", operator.user_code),
            ("生成时角色", operator.role_code),
            ("任务图斑数", summary.total_plot_count),
            ("监测总面积（公顷）", summary.total_area_ha),
            ("监测总面积（亩）", summary.total_area_mu),
            ("耕地面积（公顷）", summary.farmland_area_ha),
            ("平均图斑面积（公顷）", summary.average_plot_area_ha),
            ("作物录入完成率（%）", summary.crop_assignment_rate),
            ("生成依据", comment),
        )
        for row in summary_rows:
            summary_sheet.append(row)
        for row in summary_sheet.iter_rows(min_row=2, max_col=1):
            row[0].font = Font(bold=True, color="344B3F")
        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 75

        cls._add_group_sheet(
            workbook,
            "地级区域",
            summary.by_city,
            chart_kind="bar",
        )
        cls._add_group_sheet(workbook, "县区", summary.by_district)
        cls._add_group_sheet(
            workbook,
            "地类",
            summary.by_land_class,
            chart_kind="pie",
        )
        cls._add_group_sheet(
            workbook,
            "作物",
            summary.by_crop_type,
            chart_kind="pie",
        )
        cls._add_group_sheet(workbook, "种植模式", summary.by_planting_mode)
        cls._add_group_sheet(workbook, "权属村", summary.by_village)

        trend_sheet = workbook.create_sheet("年度趋势")
        trend_sheet.append(
            [
                "年度",
                "面积（公顷）",
                "同比（%）",
                "来源名称",
                "来源版本",
                "记录时间",
                "是否当前实时值",
            ]
        )
        cls._style_header(trend_sheet, 1, 7)
        for item in summary.annual_trend:
            trend_sheet.append(
                [
                    item.year,
                    item.area_ha,
                    item.year_over_year,
                    item.source_name,
                    item.source_version or "",
                    item.recorded_at.isoformat(),
                    "是" if item.is_current else "否",
                ]
            )
        for column, width in zip(
            "ABCDEFG",
            (12, 18, 14, 30, 20, 28, 18),
            strict=True,
        ):
            trend_sheet.column_dimensions[column].width = width
        if summary.annual_trend:
            chart = LineChart()
            chart.title = "年度监测面积变化趋势"
            chart.y_axis.title = "面积（公顷）"
            chart.x_axis.title = "年度"
            chart.add_data(
                Reference(
                    trend_sheet,
                    min_col=2,
                    min_row=1,
                    max_row=len(summary.annual_trend) + 1,
                ),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(
                    trend_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=len(summary.annual_trend) + 1,
                )
            )
            chart.height = 8
            chart.width = 16
            trend_sheet.add_chart(chart, "I2")
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _draw_wrapped_text(
        draw: ImageDraw.ImageDraw,
        text: str,
        position: tuple[int, int],
        font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
        fill: str,
        max_chars: int,
        line_height: int,
    ) -> int:
        """按中文字符宽度绘制多行文本。

        Args:
            draw: Pillow 画布。
            text: 待绘制文本。
            position: 左上角坐标。
            font: 字体。
            fill: 颜色。
            max_chars: 每行最大字符数。
            line_height: 行高。

        Returns:
            int: 绘制后的底部纵坐标。
        """
        x, y = position
        lines = [
            text[index : index + max_chars]
            for index in range(0, len(text), max_chars)
        ]
        for line in lines or [""]:
            draw.text((x, y), line, fill=fill, font=font)
            y += line_height
        return y

    @classmethod
    def _draw_bar_chart(
        cls,
        draw: ImageDraw.ImageDraw,
        items: list[AreaGroupItem],
        title: str,
        box: tuple[int, int, int, int],
    ) -> None:
        """在 PDF 页面绘制最多八项横向面积排名。

        Args:
            draw: Pillow 画布。
            items: 分组统计项。
            title: 图表标题。
            box: 图表边界。

        Returns:
            None: 原位绘图。
        """
        left, top, right, bottom = box
        title_font = cls._load_font(28)
        label_font = cls._load_font(20)
        small_font = cls._load_font(18)
        draw.rounded_rectangle(box, radius=18, fill="#FFFFFF", outline="#D9E2DD")
        draw.text((left + 24, top + 20), title, fill="#294B39", font=title_font)
        rows = items[:8]
        if not rows:
            draw.text(
                (left + 24, top + 82),
                "暂无可用统计数据",
                fill="#7B8981",
                font=label_font,
            )
            return
        max_area = max(item.area_ha for item in rows) or 1
        row_height = max((bottom - top - 100) // len(rows), 48)
        label_width = 175
        value_width = 110
        bar_left = left + label_width + 24
        bar_right = right - value_width - 24
        for index, item in enumerate(rows):
            y = top + 78 + index * row_height
            label = item.label if len(item.label) <= 9 else f"{item.label[:8]}…"
            draw.text((left + 24, y), label, fill="#3C4B43", font=small_font)
            draw.rounded_rectangle(
                (bar_left, y + 2, bar_right, y + 24),
                radius=10,
                fill="#E9EFEC",
            )
            width = max(round((bar_right - bar_left) * item.area_ha / max_area), 4)
            draw.rounded_rectangle(
                (bar_left, y + 2, bar_left + width, y + 24),
                radius=10,
                fill=REPORT_COLORS[index % len(REPORT_COLORS)],
            )
            draw.text(
                (bar_right + 12, y),
                f"{item.area_ha:.2f}",
                fill="#34473D",
                font=small_font,
            )

    @classmethod
    def _build_evidence_pages(
        cls,
        entries: list[str],
        *,
        width: int,
        height: int,
        background: str,
    ) -> list[Image.Image]:
        """把完整生成依据和年度来源按 A4 页面安全分页。

        Args:
            entries: 需要完整保留的审计文本项。
            width: 页面像素宽度。
            height: 页面像素高度。
            background: 页面背景色。

        Returns:
            list[Image.Image]: 不发生垂直溢出的证据页。
        """
        title_font = cls._load_font(40)
        body_font = cls._load_font(22)
        small_font = cls._load_font(18)
        pages: list[Image.Image] = []
        page: Image.Image | None = None
        draw: ImageDraw.ImageDraw | None = None
        y = 0

        def start_page() -> tuple[Image.Image, ImageDraw.ImageDraw, int]:
            new_page = Image.new("RGB", (width, height), background)
            new_draw = ImageDraw.Draw(new_page)
            new_draw.text(
                (70, 55),
                "生成审计与年度来源证据",
                fill="#294B39",
                font=title_font,
            )
            return new_page, new_draw, 145

        page, draw, y = start_page()
        for entry in entries:
            line_count = max(1, math.ceil(len(entry) / 46))
            required_height = line_count * 38 + 26
            if y + required_height > height - 110:
                draw.text(
                    (70, height - 65),
                    f"证据续页 · {len(pages) + 1}",
                    fill="#718078",
                    font=small_font,
                )
                pages.append(page)
                page, draw, y = start_page()
            y = cls._draw_wrapped_text(
                draw,
                entry,
                (90, y),
                body_font,
                "#46564D",
                46,
                38,
            ) + 26
        draw.text(
            (70, height - 65),
            f"证据末页 · {len(pages) + 1}",
            fill="#718078",
            font=small_font,
        )
        pages.append(page)
        return pages

    @classmethod
    def build_pdf(
        cls,
        task: object,
        summary: AreaStatisticsResponse,
        report_title: str,
        generated_at: datetime,
        operator: object,
        comment: str,
    ) -> bytes:
        """生成包含摘要、结构、排名和血缘的标准 PDF。

        Args:
            task: 当前作业任务。
            summary: 任务实时统计快照。
            report_title: 报告标题。
            generated_at: 生成时间。
            operator: 持久化项目用户。
            comment: 报告生成依据。

        Returns:
            bytes: PDF 实体字节。
        """
        width, height = 1240, 1754
        background = "#F3F6F4"
        title_font = cls._load_font(46)
        section_font = cls._load_font(30)
        body_font = cls._load_font(22)
        small_font = cls._load_font(18)

        page_one = Image.new("RGB", (width, height), background)
        draw = ImageDraw.Draw(page_one)
        draw.rectangle((0, 0, width, 250), fill="#315F45")
        cls._draw_wrapped_text(
            draw,
            report_title,
            (70, 32),
            cls._load_font(38),
            "#FFFFFF",
            28,
            48,
        )
        draw.text(
            (72, 210),
            f"{task.task_code} · {summary.monitor_year} 年 · 任务作用域正式成果",
            fill="#DDEBE3",
            font=body_font,
        )
        metrics = (
            ("任务图斑", f"{summary.total_plot_count} 块"),
            ("监测总面积", f"{summary.total_area_ha:.2f} 公顷"),
            ("折合面积", f"{summary.total_area_mu:.2f} 亩"),
            ("耕地面积", f"{summary.farmland_area_ha:.2f} 公顷"),
            ("作物录入率", f"{summary.crop_assignment_rate:.2f}%"),
            ("平均图斑", f"{summary.average_plot_area_ha:.2f} 公顷"),
        )
        for index, (label, value) in enumerate(metrics):
            column = index % 3
            row = index // 3
            left = 70 + column * 380
            top = 285 + row * 170
            draw.rounded_rectangle(
                (left, top, left + 340, top + 135),
                radius=18,
                fill="#FFFFFF",
                outline="#D9E2DD",
            )
            draw.text((left + 24, top + 24), label, fill="#718078", font=small_font)
            draw.text((left + 24, top + 65), value, fill="#2E704F", font=section_font)
        cls._draw_bar_chart(
            draw,
            summary.by_crop_type,
            "作物种植结构（公顷）",
            (70, 625, 1170, 1130),
        )
        draw.rounded_rectangle(
            (70, 1165, 1170, 1655),
            radius=18,
            fill="#FFFFFF",
            outline="#D9E2DD",
        )
        draw.text((94, 1190), "报告口径与生成审计", fill="#294B39", font=section_font)
        audit_lines = (
            f"任务名称：{task.task_name}",
            "统计范围：仅当前任务 task_plots 显式关联且未软删除的图斑。",
            f"统计时间：{summary.generated_at.isoformat()}",
            (
                f"生成人：{operator.display_name}（{operator.user_code} / "
                f"{operator.role_code}）"
            ),
            f"报告生成时间：{generated_at.isoformat()}",
            "完整生成依据与逐年度来源见后续审计证据页。",
        )
        y = 1250
        for line in audit_lines:
            y = cls._draw_wrapped_text(
                draw,
                line,
                (100, y),
                body_font,
                "#46564D",
                46,
                36,
            ) + 8

        page_two = Image.new("RGB", (width, height), background)
        draw_two = ImageDraw.Draw(page_two)
        draw_two.text(
            (70, 55),
            "多维面积结构与行政区排名",
            fill="#294B39",
            font=title_font,
        )
        cls._draw_bar_chart(
            draw_two,
            summary.by_city,
            "地级区域面积排名（公顷）",
            (70, 145, 1170, 650),
        )
        cls._draw_bar_chart(
            draw_two,
            summary.by_land_class,
            "地类面积结构（公顷）",
            (70, 680, 1170, 1160),
        )
        cls._draw_bar_chart(
            draw_two,
            summary.by_planting_mode,
            "种植模式面积结构（公顷）",
            (70, 1190, 1170, 1680),
        )

        page_three = Image.new("RGB", (width, height), background)
        draw_three = ImageDraw.Draw(page_three)
        draw_three.text(
            (70, 55),
            "年度变化趋势与来源快照",
            fill="#294B39",
            font=title_font,
        )
        chart_box = (70, 150, 1170, 820)
        draw_three.rounded_rectangle(
            chart_box,
            radius=18,
            fill="#FFFFFF",
            outline="#D9E2DD",
        )
        trend = summary.annual_trend
        if trend:
            max_area = max(item.area_ha for item in trend) or 1
            plot_left, plot_right, plot_bottom = 130, 1110, 720
            draw_three.line(
                (plot_left, plot_bottom, plot_right, plot_bottom),
                fill="#AAB7B0",
                width=3,
            )
            step = (plot_right - plot_left) / max(len(trend) - 1, 1)
            label_stride = max(1, math.ceil(len(trend) / 10))
            points: list[tuple[int, int]] = []
            for index, item in enumerate(trend):
                x = round(plot_left + index * step)
                y = round(plot_bottom - item.area_ha / max_area * 390)
                points.append((x, y))
                draw_three.ellipse((x - 8, y - 8, x + 8, y + 8), fill="#347C57")
                if index % label_stride == 0 or index == len(trend) - 1:
                    draw_three.text(
                        (x - 34, plot_bottom + 18),
                        str(item.year),
                        fill="#53635A",
                        font=small_font,
                    )
                    draw_three.text(
                        (x - 50, y - 42),
                        f"{item.area_ha:.2f}",
                        fill="#2E704F",
                        font=small_font,
                    )
            if len(points) > 1:
                draw_three.line(points, fill="#347C57", width=5)
        else:
            draw_three.text(
                (110, 250),
                "暂无真实年度统计快照",
                fill="#7B8981",
                font=body_font,
            )
        draw_three.text(
            (94, 860),
            "年度来源证据摘要",
            fill="#294B39",
            font=section_font,
        )
        draw_three.text(
            (100, 925),
            (
                f"本版本包含 {len(trend)} 个年度趋势点；"
                "完整来源、版本和记录时间见后续审计证据页。"
            ),
            fill="#46564D",
            font=body_font,
        )
        evidence_entries = [
            f"报告标题：{report_title}",
            f"任务名称：{task.task_name}",
            f"任务编号：{task.task_code}",
            "统计范围：仅当前任务 task_plots 显式关联且未软删除的图斑。",
            f"统计时间：{summary.generated_at.isoformat()}",
            (
                f"生成人：{operator.display_name}（{operator.user_code} / "
                f"{operator.role_code}）"
            ),
            f"报告生成时间：{generated_at.isoformat()}",
            f"生成依据：{comment}",
        ]
        for item in trend:
            source = item.source_name
            if item.source_version:
                source = f"{source} / {item.source_version}"
            evidence_entries.append(
                f"{item.year} 年：{item.area_ha:.2f} 公顷；"
                f"{source}；记录于 {item.recorded_at.isoformat()}"
            )
        draw_three.text(
            (70, 1680),
            "本报告由服务端依据数据库当前任务作用域生成；后续任务或历史快照变化将使本版本转为历史。",
            fill="#6E7D75",
            font=small_font,
        )

        evidence_pages = cls._build_evidence_pages(
            evidence_entries,
            width=width,
            height=height,
            background=background,
        )
        output = BytesIO()
        page_one.save(
            output,
            format="PDF",
            save_all=True,
            append_images=[page_two, page_three, *evidence_pages],
            resolution=150.0,
        )
        return output.getvalue()

    @staticmethod
    def build_manifest(
        task: object,
        summary: AreaStatisticsResponse,
        report_code: str,
        report_title: str,
        version: int,
        generated_at: datetime,
        operator: object,
        comment: str,
        xlsx_content: bytes,
        pdf_content: bytes,
        history_snapshot_count: int,
        history_latest_updated_at: datetime | None,
    ) -> dict:
        """构建报告包内逐文件校验和来源快照清单。

        Args:
            task: 当前作业任务。
            summary: 当前面积统计快照。
            report_code: 报告编号。
            report_title: 报告标题。
            version: 报告版本。
            generated_at: 生成时间。
            operator: 持久化项目用户。
            comment: 生成依据。
            xlsx_content: XLSX 字节。
            pdf_content: PDF 字节。
            history_snapshot_count: 历史快照数量。
            history_latest_updated_at: 历史快照最近更新时间。

        Returns:
            dict: 可序列化报告清单。
        """
        statistics_payload = summary.model_dump(mode="json")
        statistics_checksum = sha256(
            json.dumps(
                statistics_payload,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest()
        return {
            "schema_version": "statistics-report-v1",
            "report_code": report_code,
            "report_title": report_title,
            "version": version,
            "task": {
                "task_code": task.task_code,
                "task_name": task.task_name,
                "task_updated_at_snapshot": task.updated_at.isoformat(),
                "task_plot_count": summary.total_plot_count,
                "monitor_year": summary.monitor_year,
            },
            "statistics_snapshot": {
                "generated_at": summary.generated_at.isoformat(),
                "total_plot_count": summary.total_plot_count,
                "total_area_ha": summary.total_area_ha,
                "total_area_mu": summary.total_area_mu,
                "farmland_area_ha": summary.farmland_area_ha,
                "crop_assignment_rate": summary.crop_assignment_rate,
                "dimension_row_counts": {
                    "city": len(summary.by_city),
                    "district": len(summary.by_district),
                    "land_class": len(summary.by_land_class),
                    "crop_type": len(summary.by_crop_type),
                    "planting_mode": len(summary.by_planting_mode),
                    "village": len(summary.by_village),
                    "annual_trend": len(summary.annual_trend),
                },
                "canonical_payload_sha256": statistics_checksum,
            },
            "history_snapshot": {
                "count": history_snapshot_count,
                "latest_updated_at": (
                    history_latest_updated_at.isoformat()
                    if history_latest_updated_at
                    else None
                ),
            },
            "generator": {
                "display_name": operator.display_name,
                "user_code": operator.user_code,
                "role_code": operator.role_code,
                "generated_at": generated_at.isoformat(),
                "comment": comment,
            },
            "files": [
                {
                    "path": f"{report_code}.xlsx",
                    "format": "XLSX",
                    "file_size_bytes": len(xlsx_content),
                    "checksum_sha256": sha256(xlsx_content).hexdigest(),
                    "description": "六维面积统计、年度趋势及内嵌图表",
                },
                {
                    "path": f"{report_code}.pdf",
                    "format": "PDF",
                    "file_size_bytes": len(pdf_content),
                    "checksum_sha256": sha256(pdf_content).hexdigest(),
                    "description": "标准面积监测统计报告",
                },
            ],
        }

    @staticmethod
    def build_bundle(
        report_code: str,
        xlsx_content: bytes,
        pdf_content: bytes,
        manifest: dict,
    ) -> bytes:
        """把 XLSX、PDF 和 manifest 原子组合为 ZIP 字节。

        Args:
            report_code: 报告编号。
            xlsx_content: XLSX 字节。
            pdf_content: PDF 字节。
            manifest: 报告清单。

        Returns:
            bytes: ZIP 报告包字节。
        """
        output = BytesIO()
        with ZipFile(output, "w", ZIP_DEFLATED) as archive:
            archive.writestr(f"{report_code}.xlsx", xlsx_content)
            archive.writestr(f"{report_code}.pdf", pdf_content)
            archive.writestr(
                "manifest.json",
                json.dumps(
                    manifest,
                    ensure_ascii=False,
                    indent=2,
                ).encode("utf-8"),
            )
        return output.getvalue()
