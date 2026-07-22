"""面积统计 XLSX/PDF 正式报告包测试。"""

import asyncio
import hashlib
import json
import re
from copy import deepcopy
from datetime import UTC, datetime, timedelta
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from app.core.exceptions import PermissionDeniedException, ValidationException
from app.models.statistics_report import StatisticsReport
from app.schemas.statistics import (
    AreaGroupItem,
    AreaStatisticsResponse,
    AreaTrendItem,
    StatisticsReportGenerateRequest,
)
from app.services.statistics_report_renderer import StatisticsReportRenderer
from app.services.statistics_report_service import StatisticsReportService


def build_summary() -> AreaStatisticsResponse:
    """构造包含六个维度和真实年度来源的统计快照。

    Returns:
        AreaStatisticsResponse: 可生成正式报告的统计响应。
    """
    city = AreaGroupItem(
        code="230100",
        label="哈尔滨市",
        parent_label="黑龙江省",
        plot_count=12,
        area_ha=125.5,
        area_mu=1882.5,
        percentage=100,
    )
    district = city.model_copy(
        update={
            "code": "230109",
            "label": "松北区",
            "parent_label": "哈尔滨市",
        }
    )
    generated_at = datetime(2026, 7, 23, 8, 0, tzinfo=UTC)
    return AreaStatisticsResponse(
        task_code="RS-TEST",
        monitor_year=2026,
        generated_at=generated_at,
        total_plot_count=12,
        total_area_ha=125.5,
        total_area_mu=1882.5,
        average_plot_area_ha=10.46,
        farmland_area_ha=100.2,
        crop_assigned_plot_count=10,
        crop_assignment_rate=83.33,
        by_land_class=[city.model_copy(update={"label": "耕地"})],
        by_crop_type=[city.model_copy(update={"label": "玉米"})],
        by_planting_mode=[city.model_copy(update={"label": "单季种植"})],
        by_city=[city],
        by_district=[district],
        by_village=[city.model_copy(update={"label": "幸福村"})],
        annual_trend=[
            AreaTrendItem(
                year=2025,
                area_ha=120,
                year_over_year=None,
                source_name="2025 年农业农村统计年报",
                source_version="final-v1",
                recorded_at=datetime(2025, 12, 31, tzinfo=UTC),
                is_current=False,
            ),
            AreaTrendItem(
                year=2026,
                area_ha=125.5,
                year_over_year=4.58,
                source_name="当前任务实时汇总",
                source_version=None,
                recorded_at=generated_at,
                is_current=True,
            ),
        ],
    )


def build_task() -> SimpleNamespace:
    """构造统计报告关联任务。

    Returns:
        SimpleNamespace: 带稳定更新时间的任务。
    """
    return SimpleNamespace(
        id=7,
        project_id=3,
        task_code="RS-TEST",
        task_name="黑龙江省农业遥感监测测试任务",
        updated_at=datetime(2026, 7, 23, 7, 30, tzinfo=UTC),
    )


def build_operator() -> SimpleNamespace:
    """构造持久化项目负责人。

    Returns:
        SimpleNamespace: 稳定用户编码和角色快照。
    """
    return SimpleNamespace(
        display_name="赵志远",
        user_code="manager-zhao-zhiyuan",
        role_code="project_manager",
    )


def test_statistics_report_renderer_builds_xlsx_pdf_and_manifest() -> None:
    """验证报告实体包含六维工作表、图表、PDF 和逐文件校验。"""
    renderer = StatisticsReportRenderer()
    task = build_task()
    summary = build_summary()
    operator = build_operator()
    generated_at = datetime(2026, 7, 23, 9, 0, tzinfo=UTC)

    xlsx_content = renderer.build_xlsx(
        task,
        summary,
        "黑龙江省农作物种植面积监测统计报告",
        generated_at,
        operator,
        "测试正式报告生成",
    )
    pdf_content = renderer.build_pdf(
        task,
        summary,
        "黑龙江省农作物种植面积监测统计报告",
        generated_at,
        operator,
        "测试正式报告生成",
    )
    workbook = load_workbook(BytesIO(xlsx_content))
    assert workbook.sheetnames == [
        "报告摘要",
        "地级区域",
        "县区",
        "地类",
        "作物",
        "种植模式",
        "权属村",
        "年度趋势",
    ]
    assert workbook["地级区域"]._charts
    assert workbook["地类"]._charts
    assert workbook["作物"]._charts
    assert workbook["年度趋势"]._charts
    assert pdf_content.startswith(b"%PDF")

    report_code = "STRPT-RS-TEST-V001"
    manifest = renderer.build_manifest(
        task,
        summary,
        report_code,
        "黑龙江省农作物种植面积监测统计报告",
        1,
        generated_at,
        operator,
        "测试正式报告生成",
        xlsx_content,
        pdf_content,
        1,
        datetime(2026, 1, 1, tzinfo=UTC),
    )
    bundle = renderer.build_bundle(
        report_code,
        xlsx_content,
        pdf_content,
        manifest,
    )
    with ZipFile(BytesIO(bundle)) as archive:
        assert set(archive.namelist()) == {
            f"{report_code}.xlsx",
            f"{report_code}.pdf",
            "manifest.json",
        }
        assert json.loads(archive.read("manifest.json")) == manifest


def test_statistics_report_pdf_paginates_long_evidence() -> None:
    """验证最大生成依据和多年来源会完整分页而不是绘制到页面外。"""
    renderer = StatisticsReportRenderer()
    summary = build_summary()
    trend = [
        AreaTrendItem(
            year=1980 + index,
            area_ha=100 + index,
            year_over_year=None if index == 0 else 1,
            source_name=f"第 {index + 1} 年农业农村统计验收成果",
            source_version=f"final-v{index + 1}",
            recorded_at=datetime(1980 + index, 12, 31, tzinfo=UTC),
            is_current=index == 39,
        )
        for index in range(40)
    ]
    summary = summary.model_copy(update={"annual_trend": trend})

    pdf_content = renderer.build_pdf(
        build_task(),
        summary,
        "黑龙江省农作物种植面积监测统计成果正式报告（长期序列审计版）",
        datetime(2026, 7, 23, 9, 0, tzinfo=UTC),
        build_operator(),
        "依据" * 250,
    )

    page_count = len(re.findall(rb"/Type\s*/Page\b", pdf_content))
    assert page_count >= 5


def test_statistics_report_title_rejects_unrenderable_length() -> None:
    """验证报告标题限制在 PDF 页眉可完整呈现的 80 字以内。"""
    with pytest.raises(ValueError):
        StatisticsReportGenerateRequest(
            operator_code="manager-zhao-zhiyuan",
            report_title="黑" * 81,
            comment="验证超长标题被接口模型拒绝",
        )


def test_statistics_report_service_generates_and_verifies_bundle(
    tmp_path: Path,
) -> None:
    """验证项目负责人生成受控 ZIP 并保存任务与历史快照。"""
    task = build_task()
    summary = build_summary()
    operator = build_operator()
    report_dao = AsyncMock()
    report_dao.get_history_state.return_value = (
        1,
        datetime(2026, 1, 1, tzinfo=UTC),
    )
    report_dao.get_next_version.return_value = 1
    report_dao.supersede_completed_reports.return_value = 0

    async def add_report(_db: object, report: object) -> object:
        report.id = 11
        return report

    report_dao.add_report.side_effect = add_report
    workbench_dao = AsyncMock()
    workbench_dao.get_task_by_code_for_update.return_value = task
    statistics_service = AsyncMock()
    statistics_service.get_area_statistics.return_value = summary
    user_service = AsyncMock()
    user_service.require_capability.return_value = operator
    service = StatisticsReportService(
        dao=report_dao,
        statistics_dao=AsyncMock(),
        workbench_dao=workbench_dao,
        statistics_service=statistics_service,
        project_user_service=user_service,
        storage_root=tmp_path,
    )
    db = AsyncMock()

    response = asyncio.run(
        service.generate_report(
            db,
            task.task_code,
            StatisticsReportGenerateRequest(
                operator_code=operator.user_code,
                report_title="黑龙江省农作物种植面积监测统计报告",
                comment="依据当前任务实体图斑和真实历史快照生成",
            ),
        )
    )

    report = report_dao.add_report.await_args.args[1]
    path = service.verify_report_bundle(report)
    assert path.is_file()
    assert response.version == 1
    assert response.is_current is True
    assert response.task_plot_count == 12
    assert response.history_snapshot_count == 1
    assert response.xlsx_checksum_sha256 == report.xlsx_checksum_sha256
    assert response.pdf_checksum_sha256 == report.pdf_checksum_sha256
    user_service.require_capability.assert_awaited_once_with(
        db,
        task.project_id,
        operator.user_code,
        "generate_statistics_report",
    )
    workbench_dao.add_review_record.assert_awaited_once()
    db.commit.assert_awaited_once()

    members = service.read_verified_members(report)
    tampered_manifest = deepcopy(report.report_manifest)
    tampered_manifest["files"][0]["checksum_sha256"] = "0" * 64
    tampered_bundle = service.renderer.build_bundle(
        report.report_code,
        members[f"{report.report_code}.xlsx"],
        members[f"{report.report_code}.pdf"],
        tampered_manifest,
    )
    path.write_bytes(tampered_bundle)
    report.bundle_size_bytes = path.stat().st_size
    report.bundle_checksum_sha256 = hashlib.sha256(tampered_bundle).hexdigest()
    report.report_manifest = tampered_manifest
    with pytest.raises(ValidationException, match="XLSX 证据"):
        service.verify_report_bundle(report)


def test_statistics_report_detects_stale_task_and_corruption(tmp_path: Path) -> None:
    """验证任务变化使报告失效，实体篡改会阻止下载和归档。"""
    task = build_task()
    report = SimpleNamespace(
        status="completed",
        task_plot_count=12,
        task_updated_at_snapshot=task.updated_at,
        history_snapshot_count=1,
        history_latest_updated_at=datetime(2026, 1, 1, tzinfo=UTC),
    )
    assert StatisticsReportService.get_stale_reason(
        report,
        task,
        12,
        (1, datetime(2026, 1, 1, tzinfo=UTC)),
    ) is None
    task.updated_at += timedelta(seconds=1)
    assert "任务数据" in str(
        StatisticsReportService.get_stale_reason(
            report,
            task,
            12,
            (1, datetime(2026, 1, 1, tzinfo=UTC)),
        )
    )

    bad_path = tmp_path / "bad.zip"
    bad_path.write_bytes(b"PK\x03\x04tampered")
    invalid_report = SimpleNamespace(
        bundle_uri="storage://statistics-reports/bad.zip",
        bundle_size_bytes=bad_path.stat().st_size,
        bundle_checksum_sha256="0" * 64,
    )
    service = StatisticsReportService(storage_root=tmp_path)
    with pytest.raises(ValidationException, match="SHA-256"):
        service.verify_report_bundle(invalid_report)


def build_report_model(task: SimpleNamespace) -> StatisticsReport:
    """构造列表与权限测试所需的完整报告模型。

    Args:
        task: 报告关联任务。

    Returns:
        StatisticsReport: 无需真实实体文件的报告摘要。
    """
    return StatisticsReport(
        id=11,
        task_id=task.id,
        report_code="STRPT-RS-TEST-V001",
        report_title="黑龙江省农作物种植面积监测统计报告",
        version=1,
        status="completed",
        bundle_uri="storage://statistics-reports/RS-TEST/report.zip",
        bundle_size_bytes=10,
        bundle_checksum_sha256="a" * 64,
        xlsx_size_bytes=4,
        xlsx_checksum_sha256="b" * 64,
        pdf_size_bytes=4,
        pdf_checksum_sha256="c" * 64,
        task_plot_count=12,
        task_updated_at_snapshot=task.updated_at,
        history_snapshot_count=1,
        history_latest_updated_at=datetime(2026, 1, 1, tzinfo=UTC),
        report_manifest={"schema_version": "statistics-report-v1"},
        generation_comment="测试报告",
        generated_by="赵志远",
        generated_by_code="manager-zhao-zhiyuan",
        generated_by_role="project_manager",
        generated_at=datetime(2026, 7, 23, 9, 0, tzinfo=UTC),
    )


def test_statistics_report_list_is_read_only_and_does_not_scan_files() -> None:
    """验证列表只计算来源状态，不读取所有历史 ZIP 或在 GET 中提交。"""
    task = build_task()
    report = build_report_model(task)
    task.updated_at += timedelta(seconds=1)
    report_dao = AsyncMock()
    report_dao.get_history_state.return_value = (
        1,
        datetime(2026, 1, 1, tzinfo=UTC),
    )
    report_dao.list_reports.return_value = [report]
    statistics_dao = AsyncMock()
    statistics_dao.get_totals.return_value = (12, 100)
    workbench_dao = AsyncMock()
    workbench_dao.get_task_by_code.return_value = task
    service = StatisticsReportService(
        dao=report_dao,
        statistics_dao=statistics_dao,
        workbench_dao=workbench_dao,
    )
    service.verify_report_bundle = MagicMock(
        side_effect=AssertionError("列表不应扫描报告实体")
    )
    db = AsyncMock()

    response = asyncio.run(service.list_reports(db, task.task_code))

    assert response.items[0].status == "superseded"
    assert response.items[0].is_current is False
    service.verify_report_bundle.assert_not_called()
    db.commit.assert_not_awaited()


def test_statistics_report_download_checks_persisted_capability_first() -> None:
    """验证无下载能力的用户不能触发报告实体读取。"""
    task = build_task()
    report = build_report_model(task)
    report_dao = AsyncMock()
    report_dao.get_report_by_code.return_value = report
    workbench_dao = AsyncMock()
    workbench_dao.get_task_by_id.return_value = task
    user_service = AsyncMock()
    user_service.require_capability.side_effect = PermissionDeniedException(
        "内业解译员无权下载正式统计报告"
    )
    service = StatisticsReportService(
        dao=report_dao,
        workbench_dao=workbench_dao,
        project_user_service=user_service,
    )
    service.verify_report_bundle = MagicMock()

    with pytest.raises(PermissionDeniedException):
        asyncio.run(
            service.authorize_download(
                AsyncMock(),
                report.report_code,
                "interp-li-jing",
            )
        )

    service.verify_report_bundle.assert_not_called()
