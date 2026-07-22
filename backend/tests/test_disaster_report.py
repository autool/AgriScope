"""灾害监测专题报告实体生成、门禁和完整性测试。"""

import asyncio
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook

from app.core.exceptions import ValidationException
from app.schemas.disaster import (
    DisasterGroupItem,
    DisasterPatchResponse,
    DisasterReportGenerateRequest,
    DisasterSummaryResponse,
)
from app.services.disaster_report_service import DisasterReportService


def build_summary(*, pending: bool = False) -> DisasterSummaryResponse:
    """构造含真实 Polygon 结构的灾害汇总。

    Args:
        pending: 是否保留待复核斑块。

    Returns:
        DisasterSummaryResponse: 可用于报告渲染的汇总。
    """
    item = DisasterPatchResponse(
        patch_code="TEST-DS-001",
        disaster_type="洪涝",
        severity="重度",
        affected_area_ha=12.5,
        crop_type="玉米",
        detected_at=date(2026, 7, 20),
        ndvi_change=-0.4,
        status="pending" if pending else "confirmed",
        source="公开灾害模型",
        source_uri="https://example.test/disaster.geojson",
        source_version="20260720",
        source_feature_id="feature-001",
        source_checksum_sha256="a" * 64,
        import_batch_code="DSIMP-001",
        imported_by="李静",
        imported_by_code="interpreter-li-jing",
        imported_by_role="interpreter",
        reviewed_by=None if pending else "王海峰",
        reviewed_by_code=None if pending else "quality-wang-haifeng",
        reviewed_by_role=None if pending else "quality_inspector",
        review_comment=None if pending else "现场与影像共同确认",
        reviewed_at=None if pending else datetime.now(UTC),
        geometry={
            "type": "Polygon",
            "coordinates": [
                [
                    [126.5, 45.7],
                    [126.6, 45.7],
                    [126.6, 45.8],
                    [126.5, 45.8],
                    [126.5, 45.7],
                ]
            ],
        },
    )
    group = DisasterGroupItem(
        label="重度",
        patch_count=1,
        area_ha=12.5,
        percentage=100,
    )
    return DisasterSummaryResponse(
        task_code="RS-2026-045",
        generated_at=datetime.now(UTC),
        total_patches=1,
        affected_area_ha=12.5,
        pending_count=1 if pending else 0,
        confirmed_count=0 if pending else 1,
        by_severity=[group],
        by_type=[
            DisasterGroupItem(
                label="洪涝",
                patch_count=1,
                area_ha=12.5,
                percentage=100,
            )
        ],
        items=[item],
        feature_collection={"type": "FeatureCollection", "features": []},
    )


def build_service(
    tmp_path: Path,
    *,
    pending: bool = False,
) -> tuple[DisasterReportService, AsyncMock, AsyncMock]:
    """构造具备稳定用户和斑块快照的报告服务。

    Args:
        tmp_path: pytest 临时目录。
        pending: 是否使用待复核斑块。

    Returns:
        tuple: 报告服务、报告 DAO 和工作台 DAO。
    """
    report_dao = AsyncMock()
    report_dao.supersede_completed_reports.return_value = 0
    report_dao.add_report.side_effect = lambda _db, report: report
    disaster_dao = AsyncMock()
    patch = SimpleNamespace(
        status="pending" if pending else "confirmed",
        updated_at=datetime(2026, 7, 23, 1, 0, tzinfo=UTC),
    )
    disaster_dao.get_patches.return_value = [(patch, "{}")]
    workbench_dao = AsyncMock()
    workbench_dao.get_task_by_code_for_update.return_value = SimpleNamespace(
        id=1,
        project_id=1,
        task_code="RS-2026-045",
        task_name="黑龙江省农业遥感监测任务",
        administrative_region="黑龙江省",
        status="interpreting",
        updated_at=datetime.now(UTC),
    )
    user_service = AsyncMock()
    user_service.require_capability.return_value = SimpleNamespace(
        display_name="王海峰",
        user_code="quality-wang-haifeng",
        role_code="quality_inspector",
    )
    disaster_service = AsyncMock()
    disaster_service.get_summary.return_value = build_summary(pending=pending)
    service = DisasterReportService(
        dao=report_dao,
        disaster_dao=disaster_dao,
        workbench_dao=workbench_dao,
        project_user_service=user_service,
        disaster_service=disaster_service,
        storage_root=tmp_path,
    )
    return service, report_dao, workbench_dao


def test_report_workbook_contains_map_charts_details_and_lineage() -> None:
    """验证报告实体包含摘要、图表、分布图、明细和来源审计。"""
    task = SimpleNamespace(
        task_code="RS-2026-045",
        task_name="黑龙江省农业遥感监测任务",
        administrative_region="黑龙江省",
    )
    generator = SimpleNamespace(
        display_name="王海峰",
        user_code="quality-wang-haifeng",
        role_code="quality_inspector",
    )
    content, manifest = DisasterReportService._build_workbook(
        task,
        build_summary(),
        "农业灾害遥感监测专题报告",
        datetime.now(UTC),
        generator,
        "全部灾害斑块已完成人工复核",
    )

    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["报告摘要", "等级统计", "类型统计", "斑块明细"]
    assert workbook["报告摘要"]._images
    assert workbook["等级统计"]._charts
    assert workbook["类型统计"]._charts
    assert workbook["斑块明细"]["A2"].value == "TEST-DS-001"
    assert manifest["source_checksums"] == ["a" * 64]
    assert manifest["map_png_size_bytes"] > 0


def test_generate_report_blocks_pending_patches(tmp_path: Path) -> None:
    """验证仍有待复核灾害斑块时不能生成专题报告。"""
    service, report_dao, _ = build_service(tmp_path, pending=True)

    with pytest.raises(ValidationException, match="待复核"):
        asyncio.run(
            service.generate_report(
                AsyncMock(),
                "RS-2026-045",
                DisasterReportGenerateRequest(
                    operator_code="quality-wang-haifeng",
                    report_title="农业灾害遥感监测专题报告",
                    comment="等待全部灾害斑块完成复核后生成",
                ),
            )
        )

    report_dao.add_report.assert_not_awaited()
    assert not list(tmp_path.rglob("*.xlsx"))


def test_generate_report_publishes_checksum_backed_xlsx(tmp_path: Path) -> None:
    """验证报告原子发布并保存实体大小、SHA-256 和稳定用户审计。"""
    service, report_dao, workbench_dao = build_service(tmp_path)

    response = asyncio.run(
        service.generate_report(
            AsyncMock(),
            "RS-2026-045",
            DisasterReportGenerateRequest(
                operator_code="quality-wang-haifeng",
                report_title="农业灾害遥感监测专题报告",
                comment="全部灾害斑块已完成人工复核并用于成果审核",
            ),
        )
    )

    report = report_dao.add_report.await_args.args[1]
    path = service.verify_report_file(report)
    assert path.read_bytes().startswith(b"PK\x03\x04")
    assert response.status == "completed"
    assert response.is_current is True
    assert response.file_size_bytes == path.stat().st_size
    assert len(response.checksum_sha256) == 64
    review = workbench_dao.add_review_record.await_args.args[1]
    assert review.action == "disaster_report_generated"


def test_verify_report_rejects_tampered_entity(tmp_path: Path) -> None:
    """验证下载前能够拒绝被篡改的灾害专题报告。"""
    service, report_dao, _ = build_service(tmp_path)
    asyncio.run(
        service.generate_report(
            AsyncMock(),
            "RS-2026-045",
            DisasterReportGenerateRequest(
                operator_code="quality-wang-haifeng",
                report_title="农业灾害遥感监测专题报告",
                comment="全部灾害斑块已完成人工复核并用于成果审核",
            ),
        )
    )
    report = report_dao.add_report.await_args.args[1]
    path = service.verify_report_file(report)
    path.write_bytes(path.read_bytes() + b"tampered")

    with pytest.raises(ValidationException, match="大小校验失败"):
        service.verify_report_file(report)
