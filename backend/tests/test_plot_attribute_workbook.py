"""任务地块属性 Excel 原子导入导出测试。"""

import asyncio
from datetime import UTC, datetime
from hashlib import sha256
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import Workbook, load_workbook

from app.core.exceptions import PermissionDeniedException, ValidationException
from app.models.plot_attribute_workbook import PlotAttributeImportBatch
from app.schemas.plot_attribute_workbook import (
    PlotAttributeWorkbookExportRequest,
    PlotAttributeWorkbookImportMetadata,
)
from app.services.plot_attribute_workbook_parser import (
    WORKBOOK_HEADERS,
    PlotAttributeWorkbookParser,
)
from app.services.plot_attribute_workbook_service import (
    PlotAttributeWorkbookService,
)
from app.services.plot_attribute_workbook_storage import (
    PlotAttributeWorkbookStorage,
    StoredPlotAttributeWorkbook,
)


def build_workbook_bytes(rows: list[list[object]]) -> bytes:
    """构造标准地块属性 XLSX。

    Args:
        rows: 表头之后的工作簿数据行。

    Returns:
        bytes: XLSX 实体字节。
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "地块属性"
    worksheet.append(list(WORKBOOK_HEADERS))
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def build_plot(
    plot_code: str,
    *,
    version: int = 1,
    owner_village: str = "原权属村",
    land_class: str = "耕地",
    crop_type: str | None = "玉米",
) -> SimpleNamespace:
    """构造可锁定更新的任务图斑。

    Args:
        plot_code: 图斑业务编号。
        version: 当前版本。
        owner_village: 当前权属村。
        land_class: 当前地类。
        crop_type: 当前作物。

    Returns:
        SimpleNamespace: 满足服务属性访问的图斑替身。
    """
    return SimpleNamespace(
        plot_code=plot_code,
        version=version,
        owner_village=owner_village,
        land_class=land_class,
        crop_type=crop_type,
        planting_mode="单季种植" if land_class == "耕地" else None,
        irrigation_condition="良好" if land_class == "耕地" else None,
        interpretation_status="interpreted",
        geom=f"geometry-{plot_code}",
        updated_at=datetime.now(UTC),
    )


def build_service(
    plots: list[SimpleNamespace],
) -> tuple[
    PlotAttributeWorkbookService,
    AsyncMock,
    AsyncMock,
    AsyncMock,
    MagicMock,
]:
    """构造带任务、权限和受控存储替身的服务。

    Args:
        plots: 任务内有效图斑列表。

    Returns:
        tuple: 服务、批次 DAO、工作台 DAO、用户服务和存储替身。
    """
    task = SimpleNamespace(
        id=7,
        project_id=3,
        status="interpreting",
        total_plots=len(plots),
        completed_plots=len(plots),
        quality_score=96,
        updated_at=datetime.now(UTC),
    )
    dao = AsyncMock()

    async def add_batch(_db, batch):
        batch.id = 1
        return batch

    dao.add_batch.side_effect = add_batch
    dao.list_batches.return_value = []
    workbench_dao = AsyncMock()
    workbench_dao.get_task_by_code.return_value = task
    workbench_dao.get_task_by_code_for_update.return_value = task
    workbench_dao.get_task_plots.return_value = plots
    workbench_dao.get_task_plots_by_codes.return_value = plots
    workbench_dao.get_task_plots_by_codes_for_update.return_value = plots
    workbench_dao.count_plot_progress.return_value = (len(plots), len(plots))
    user_service = AsyncMock()
    user_service.require_capability.return_value = SimpleNamespace(
        user_code="interp-li-jing",
        display_name="李静",
        role_code="interpreter",
    )
    storage = MagicMock()

    def store(filename: str, content: bytes) -> StoredPlotAttributeWorkbook:
        checksum = sha256(content).hexdigest()
        return StoredPlotAttributeWorkbook(
            path=SimpleNamespace(unlink=MagicMock()),
            file_uri=f"storage://plot-attribute-imports/{checksum}.xlsx",
            file_size_bytes=len(content),
            checksum_sha256=checksum,
            created_new=False,
        )

    storage.store.side_effect = store
    service = PlotAttributeWorkbookService(
        dao=dao,
        workbench_dao=workbench_dao,
        storage=storage,
        user_service=user_service,
    )
    return service, dao, workbench_dao, user_service, storage


def build_metadata() -> PlotAttributeWorkbookImportMetadata:
    """构造标准导入元数据。

    Returns:
        PlotAttributeWorkbookImportMetadata: 稳定操作人和判读依据。
    """
    return PlotAttributeWorkbookImportMetadata(
        operator_code="interp-li-jing",
        comment="依据 2026-07-20 Sentinel-2 影像与外业调查表逐图斑核定",
    )


def test_parser_accepts_mixed_per_row_attributes() -> None:
    """验证同一工作簿可为每个图斑填写不同属性。"""
    parser = PlotAttributeWorkbookParser()
    content = build_workbook_bytes(
        [
            ["PLOT-001", 2, "幸福村", "耕地", "大豆", "轮作", "良好"],
            ["PLOT-002", 4, "建设村", "建设用地", "", "", ""],
        ]
    )

    rows = parser.parse("plot-attributes.xlsx", content)

    assert len(rows) == 2
    assert rows[0].crop_type == "大豆"
    assert rows[1].land_class == "建设用地"
    assert rows[1].crop_type is None


def test_parser_rejects_formula_cells() -> None:
    """验证任一公式单元格都会拒绝整个工作簿。"""
    parser = PlotAttributeWorkbookParser()
    content = build_workbook_bytes(
        [["PLOT-001", 1, "幸福村", "耕地", '="玉"&"米"', "单季种植", "良好"]]
    )

    with pytest.raises(ValidationException, match="包含公式"):
        parser.parse("plot-attributes.xlsx", content)


def test_parser_rejects_formula_in_non_business_sheet() -> None:
    """验证说明页等非业务工作表中的公式也不能绕过安全校验。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "地块属性"
    worksheet.append(list(WORKBOOK_HEADERS))
    worksheet.append(
        ["PLOT-001", 1, "幸福村", "耕地", "玉米", "单季种植", "良好"]
    )
    instructions = workbook.create_sheet("填写说明")
    instructions["A1"] = '=HYPERLINK("https://example.invalid","说明")'
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(ValidationException, match="填写说明.*包含公式"):
        PlotAttributeWorkbookParser().parse(
            "plot-attributes.xlsx",
            output.getvalue(),
        )


def test_parser_rejects_duplicate_plot_code() -> None:
    """验证批次内重复图斑编号在进入事务前被拒绝。"""
    parser = PlotAttributeWorkbookParser()
    content = build_workbook_bytes(
        [
            ["PLOT-001", 1, "幸福村", "耕地", "玉米", "单季种植", "良好"],
            ["PLOT-001", 1, "幸福村", "耕地", "大豆", "轮作", "良好"],
        ]
    )

    with pytest.raises(ValidationException, match="图斑编号 PLOT-001 重复"):
        parser.parse("plot-attributes.xlsx", content)


def test_parser_rejects_invalid_farmland_crop_logic() -> None:
    """验证耕地缺少作物或非耕地保留作物都会整批拒绝。"""
    parser = PlotAttributeWorkbookParser()
    content = build_workbook_bytes(
        [["PLOT-001", 1, "幸福村", "耕地", "", "单季种植", "良好"]]
    )

    with pytest.raises(ValidationException, match="耕地图斑必须填写作物类型"):
        parser.parse("plot-attributes.xlsx", content)


def test_export_contains_current_values_and_expected_version() -> None:
    """验证服务端导出真实当前属性和并发版本。"""
    plots = [build_plot("PLOT-001", version=7, owner_village="幸福村")]
    service, _, _, user_service, _ = build_service(plots)
    db = AsyncMock()

    result = asyncio.run(
        service.export_workbook(
            db,
            "RS-2026-045",
            PlotAttributeWorkbookExportRequest(
                operator_code="interp-li-jing",
                plot_codes=["PLOT-001"],
            ),
        )
    )

    workbook = load_workbook(BytesIO(result.content), read_only=True)
    worksheet = workbook["地块属性"]
    values = list(worksheet.values)
    workbook.close()
    assert values[0] == WORKBOOK_HEADERS
    assert values[1][:4] == ("PLOT-001", 7, "幸福村", "耕地")
    assert result.row_count == 1
    user_service.require_capability.assert_awaited_once()


def test_export_rejects_large_task_before_loading_plot_geometry() -> None:
    """验证任务超过上限时先计数拒绝，不读取数万条完整图斑几何。"""
    service, _, workbench_dao, _, _ = build_service([build_plot("PLOT-001")])
    workbench_dao.count_plot_progress.return_value = (35020, 0)

    with pytest.raises(ValidationException, match="35020.*最多导出 500"):
        asyncio.run(
            service.export_workbook(
                AsyncMock(),
                "RS-2026-045",
                PlotAttributeWorkbookExportRequest(
                    operator_code="interp-li-jing",
                ),
            )
        )

    workbench_dao.get_task_plots.assert_not_awaited()


def test_import_updates_mixed_rows_and_creates_complete_versions() -> None:
    """验证逐行差异属性一次提交并为实际变化图斑生成完整版本。"""
    first = build_plot("PLOT-001", version=2)
    second = build_plot(
        "PLOT-002",
        version=4,
        land_class="建设用地",
        crop_type=None,
    )
    service, dao, workbench_dao, _, _ = build_service([first, second])
    content = build_workbook_bytes(
        [
            ["PLOT-001", 2, "幸福村", "耕地", "大豆", "轮作", "一般"],
            ["PLOT-002", 4, "建设村", "建设用地", "", "", ""],
        ]
    )
    db = AsyncMock()

    response = asyncio.run(
        service.import_workbook(
            db,
            "RS-2026-045",
            build_metadata(),
            "plot-attributes.xlsx",
            content,
        )
    )

    assert response.row_count == 2
    assert response.changed_count == 2
    assert response.unchanged_count == 0
    assert response.checksum_sha256 == sha256(content).hexdigest()
    assert first.version == 3
    assert first.owner_village == "幸福村"
    assert first.crop_type == "大豆"
    assert second.version == 5
    versions = workbench_dao.add_plot_versions.await_args.args[1]
    assert [version.owner_village for version in versions] == ["幸福村", "建设村"]
    dao.reset_quality_evidence.assert_awaited_once_with(
        db,
        7,
        ["PLOT-001", "PLOT-002"],
    )
    workbench_dao.supersede_reverted_plot_operations.assert_awaited_once()
    db.commit.assert_awaited_once()


def test_import_rejects_stale_version_before_any_mutation() -> None:
    """验证后置行版本过期时前置图斑也不会被内存或数据库更新。"""
    first = build_plot("PLOT-001", version=2)
    second = build_plot("PLOT-002", version=5)
    service, dao, workbench_dao, _, _ = build_service([first, second])
    content = build_workbook_bytes(
        [
            ["PLOT-001", 2, "新村", "耕地", "大豆", "轮作", "一般"],
            ["PLOT-002", 4, "新村", "耕地", "大豆", "轮作", "一般"],
        ]
    )
    db = AsyncMock()

    with pytest.raises(ValidationException, match="版本已变化"):
        asyncio.run(
            service.import_workbook(
                db,
                "RS-2026-045",
                build_metadata(),
                "plot-attributes.xlsx",
                content,
            )
        )

    assert first.version == 2
    assert first.owner_village == "原权属村"
    workbench_dao.add_plot_versions.assert_not_awaited()
    dao.add_batch.assert_not_awaited()
    db.commit.assert_not_awaited()


def test_import_rejects_plot_outside_task_scope() -> None:
    """验证未知、跨任务或已删除图斑会拒绝整批。"""
    first = build_plot("PLOT-001", version=2)
    service, dao, workbench_dao, _, _ = build_service([first])
    workbench_dao.get_task_plots_by_codes_for_update.return_value = [first]
    content = build_workbook_bytes(
        [
            ["PLOT-001", 2, "幸福村", "耕地", "玉米", "单季种植", "良好"],
            ["OUTSIDE-001", 1, "越权村", "耕地", "大豆", "轮作", "一般"],
        ]
    )

    with pytest.raises(ValidationException, match="不属于任务或已删除"):
        asyncio.run(
            service.import_workbook(
                AsyncMock(),
                "RS-2026-045",
                build_metadata(),
                "plot-attributes.xlsx",
                content,
            )
        )

    dao.add_batch.assert_not_awaited()


def test_import_permission_denial_stops_before_parsing_or_storage() -> None:
    """验证无导入能力的稳定项目用户不能消耗解析或存储资源。"""
    service, _, _, user_service, storage = build_service([build_plot("PLOT-001")])
    user_service.require_capability.side_effect = PermissionDeniedException(
        "当前角色无权导入地块属性"
    )
    service.parser = MagicMock()

    with pytest.raises(PermissionDeniedException, match="无权导入"):
        asyncio.run(
            service.import_workbook(
                AsyncMock(),
                "RS-2026-045",
                build_metadata(),
                "plot-attributes.xlsx",
                b"not-read",
            )
        )

    service.parser.parse.assert_not_called()
    storage.store.assert_not_called()


def test_storage_persists_exact_size_and_sha256(tmp_path) -> None:
    """验证受控文件名、实体大小和服务端 SHA256 与原始工作簿一致。"""
    content = build_workbook_bytes(
        [["PLOT-001", 1, "幸福村", "耕地", "玉米", "单季种植", "良好"]]
    )
    storage = PlotAttributeWorkbookStorage(tmp_path)

    stored = storage.store("plot-attributes.xlsx", content)

    assert stored.created_new is True
    assert stored.file_size_bytes == len(content)
    assert stored.checksum_sha256 == sha256(content).hexdigest()
    assert stored.path.read_bytes() == content
    assert stored.file_uri.endswith(f"{stored.checksum_sha256}.xlsx")


def test_delivery_verification_reopens_workbook_and_checks_batch_evidence(
    tmp_path,
) -> None:
    """验证成果归档前重新解析 XLSX 并核对大小、SHA256 和行数。"""
    content = build_workbook_bytes(
        [["PLOT-001", 1, "幸福村", "耕地", "玉米", "单季种植", "良好"]]
    )
    storage = PlotAttributeWorkbookStorage(tmp_path)
    stored = storage.store("plot-attributes.xlsx", content)
    batch = PlotAttributeImportBatch(
        task_id=7,
        batch_code="PATTR-TEST-001",
        original_filename="plot-attributes.xlsx",
        file_uri=stored.file_uri,
        file_size_bytes=stored.file_size_bytes,
        checksum_sha256=stored.checksum_sha256,
        row_count=1,
        changed_count=1,
        unchanged_count=0,
        imported_by="李静",
        imported_by_code="interp-li-jing",
        imported_by_role="interpreter",
        import_comment="测试成果归档复核",
        imported_at=datetime.now(UTC),
    )

    verified = storage.verify_import_workbook(
        batch,
        PlotAttributeWorkbookParser(),
    )

    assert verified.path == stored.path
    assert verified.row_count == 1
    assert verified.changed_count == 1
    assert verified.checksum_sha256 == sha256(content).hexdigest()


def test_delivery_verification_rejects_tampered_workbook(tmp_path) -> None:
    """验证受控 XLSX 被修改后不能进入成果交付包。"""
    content = build_workbook_bytes(
        [["PLOT-001", 1, "幸福村", "耕地", "玉米", "单季种植", "良好"]]
    )
    storage = PlotAttributeWorkbookStorage(tmp_path)
    stored = storage.store("plot-attributes.xlsx", content)
    batch = PlotAttributeImportBatch(
        task_id=7,
        batch_code="PATTR-TEST-002",
        original_filename="plot-attributes.xlsx",
        file_uri=stored.file_uri,
        file_size_bytes=stored.file_size_bytes,
        checksum_sha256=stored.checksum_sha256,
        row_count=1,
        changed_count=0,
        unchanged_count=1,
        imported_by="李静",
        imported_by_code="interp-li-jing",
        imported_by_role="interpreter",
        import_comment="篡改检测",
        imported_at=datetime.now(UTC),
    )
    stored.path.write_bytes(content + b"tampered")

    with pytest.raises(ValidationException, match="文件大小校验失败"):
        storage.verify_import_workbook(batch, PlotAttributeWorkbookParser())
